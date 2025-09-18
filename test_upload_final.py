import os
import sys
from openpyxl import Workbook
import io

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_upload_final.db'

# Importar e testar a aplicação
from app import create_app

def test_upload_final():
    """Teste final do upload de planilha com dados completos"""
    app = create_app()
    
    with app.test_client() as client:
        print("🧪 TESTE FINAL - UPLOAD DE PLANILHA COMPLETO")
        print("=" * 60)
        
        # Criar planilha de teste com dados mais realistas
        wb = Workbook()
        ws = wb.active
        
        # Cabeçalhos
        ws['A1'] = 'nome_empreendimento'
        ws['B1'] = 'nome_empresa'
        ws['C1'] = 'cep'
        ws['D1'] = 'endereco'
        ws['E1'] = 'observacao'
        ws['F1'] = 'numero_unidade'
        ws['G1'] = 'tamanho_m2'
        ws['H1'] = 'preco_venda'
        ws['I1'] = 'mecanismo_pagamento'
        
        # Dados realistas
        dados = [
            ['Residencial Vista Verde', 'Construtora Verde Ltda', '70000-100', 'Rua das Flores, 100 - Asa Norte', 'Empreendimento sustentável', '101', 65.5, 280000.00, 'financiamento'],
            ['Residencial Vista Verde', 'Construtora Verde Ltda', '70000-100', 'Rua das Flores, 100 - Asa Norte', 'Empreendimento sustentável', '102', 58.0, 260000.00, 'à vista'],
            ['Residencial Vista Verde', 'Construtora Verde Ltda', '70000-100', 'Rua das Flores, 100 - Asa Norte', 'Empreendimento sustentável', '103', 72.0, 320000.00, 'consórcio'],
            ['Condomínio Águas Claras', 'Incorporadora Águas Ltda', '71900-000', 'Av. das Araucárias, 500 - Águas Claras', 'Próximo ao metrô', '201', 80.0, 450000.00, 'financiamento'],
            ['Condomínio Águas Claras', 'Incorporadora Águas Ltda', '71900-000', 'Av. das Araucárias, 500 - Águas Claras', 'Próximo ao metrô', '202', 85.5, 480000.00, 'Parcelamento direto 60x'],
        ]
        
        for i, linha in enumerate(dados, start=2):
            for j, valor in enumerate(linha, start=1):
                ws.cell(row=i, column=j, value=valor)
        
        print("📝 Planilha criada com:")
        print("   - 2 empreendimentos diferentes")
        print("   - 5 unidades com dados realistas")
        print("   - Diferentes formas de pagamento")
        print("   - Campo 'outros' personalizado")
        
        # Teste 1: Upload
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        print("\n1. 📤 Fazendo Upload...")
        response = client.post('/empreendimentos/upload',
                             data={'file': (file_stream, 'teste_final.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        
        if response.status_code == 200:
            data = response.get_json()
            print(f"   ✅ Empreendimentos processados: {data.get('empreendimentos_processados')}")
            print(f"   ✅ Unidades processadas: {data.get('unidades_processadas')}")
            print(f"   ✅ Erros: {data.get('erros')}")
            
            if data.get('detalhes_erros'):
                print("   ❌ Detalhes dos erros:")
                for erro in data.get('detalhes_erros'):
                    print(f"      - {erro}")
            
            # Mostrar empreendimentos criados
            if data.get('empreendimentos'):
                print("\n   📋 Empreendimentos criados:")
                for emp in data.get('empreendimentos'):
                    print(f"      - ID: {emp.get('id')} | {emp.get('nome')} | {emp.get('nome_empresa')}")
            
            # Mostrar unidades criadas
            if data.get('unidades'):
                print("\n   🏠 Unidades criadas:")
                for unidade in data.get('unidades'):
                    print(f"      - Unidade {unidade.get('numero_unidade')} | {unidade.get('tamanho_m2')}m² | R$ {unidade.get('preco_venda'):,.2f} | {unidade.get('mecanismo_pagamento')}")
        else:
            print(f"   ❌ Erro: {response.get_json()}")
            return False
        
        # Teste 2: Verificar dados no banco
        print("\n2. 🗄️ Verificando Dados no Banco...")
        
        response = client.get('/empreendimentos')
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   📊 Total de empreendimentos: {len(empreendimentos)}")
            
            for emp in empreendimentos:
                print(f"      - {emp.get('nome')} ({emp.get('nome_empresa')}) - CEP: {emp.get('cep')}")
        
        response = client.get('/api/unidades')
        if response.status_code == 200:
            unidades = response.get_json()
            print(f"   📊 Total de unidades: {len(unidades)}")
            
            # Verificar se o campo "outros" foi salvo corretamente
            for unidade in unidades:
                if 'Parcelamento direto' in unidade.get('mecanismo_pagamento', ''):
                    print(f"      ✅ Campo 'Outros' funcionando: Unidade {unidade.get('numero_unidade')} - {unidade.get('mecanismo_pagamento')}")
        
        # Teste 3: Testar com planilha com erro
        print("\n3. ❌ Testando Planilha com Erro...")
        
        wb_erro = Workbook()
        ws_erro = wb_erro.active
        
        # Cabeçalhos corretos
        ws_erro['A1'] = 'nome_empreendimento'
        ws_erro['B1'] = 'nome_empresa'
        ws_erro['C1'] = 'cep'
        
        # Dados com erro (sem CEP)
        ws_erro['A2'] = 'Empreendimento Sem CEP'
        ws_erro['B2'] = 'Construtora Erro'
        ws_erro['C2'] = ''  # CEP vazio - deve gerar erro
        
        file_stream_erro = io.BytesIO()
        wb_erro.save(file_stream_erro)
        file_stream_erro.seek(0)
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (file_stream_erro, 'teste_erro.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        
        if response.status_code == 200:
            data_erro = response.get_json()
            print(f"   Empreendimentos processados: {data_erro.get('empreendimentos_processados')}")
            print(f"   Erros encontrados: {data_erro.get('erros')}")
            
            if data_erro.get('detalhes_erros'):
                print("   ✅ Erros detectados corretamente:")
                for erro in data_erro.get('detalhes_erros'):
                    print(f"      - {erro}")
        
        print("\n" + "=" * 60)
        print("🎉 TESTE FINAL CONCLUÍDO!")
        
        # Resumo final
        print("\n📊 RESUMO FINAL:")
        print("✅ Upload de planilha funcionando corretamente")
        print("✅ Processamento de múltiplos empreendimentos")
        print("✅ Criação de unidades com dados completos")
        print("✅ Campo 'outros' personalizado funcionando")
        print("✅ Detecção e relatório de erros")
        print("✅ Interface melhorada com detalhes dos dados criados")
        
        return True

if __name__ == '__main__':
    try:
        success = test_upload_final()
        if success:
            print("\n🎯 UPLOAD DE PLANILHA TOTALMENTE FUNCIONAL!")
        else:
            print("\n❌ Problemas encontrados no upload")
            sys.exit(1)
    except Exception as e:
        print(f"\nErro: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
