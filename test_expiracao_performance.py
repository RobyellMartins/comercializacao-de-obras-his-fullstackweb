import os
import sys
import io
import json
from datetime import datetime, timedelta
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_expiracao_obras_his.db'

# Importar e testar a aplica√ß√£o
from app import create_app

def test_expiracao_automatica():
    """Testa sistema de expira√ß√£o autom√°tica"""
    app = create_app()
    
    with app.test_client() as client:
        print("‚è∞ Testando sistema de expira√ß√£o autom√°tica...")
        
        # Criar empreendimento
        print("\n1. Criando empreendimento para teste de expira√ß√£o...")
        empreendimento_teste = {
            'nome': 'Residencial Expira√ß√£o Teste',
            'nome_empresa': 'Empresa Expira√ß√£o',
            'cep': '70000-777',
            'endereco': 'Rua Expira√ß√£o, 777'
        }
        
        response = client.post('/empreendimentos',
                             data=json.dumps(empreendimento_teste),
                             content_type='application/json')
        
        if response.status_code != 201:
            print(f"   Erro ao criar empreendimento: {response.get_json()}")
            return
        
        empreendimento_id = response.get_json().get('id')
        print(f"   Empreendimento criado com ID: {empreendimento_id}")
        
        # Publicar empreendimento
        print("\n2. Publicando empreendimento...")
        response = client.post(f'/empreendimentos/{empreendimento_id}/publicar')
        print(f"   Status: {response.status_code}")
        
        if response.status_code == 200:
            dados_publicacao = response.get_json()
            publicado_em = dados_publicacao.get('publicado_em')
            expira_em = dados_publicacao.get('expira_em')
            
            print(f"   Publicado em: {publicado_em}")
            print(f"   Expira em: {expira_em}")
            
            # Verificar se a data de expira√ß√£o est√° correta (30 dias)
            if publicado_em and expira_em:
                pub_date = datetime.fromisoformat(publicado_em.replace('Z', '+00:00'))
                exp_date = datetime.fromisoformat(expira_em.replace('Z', '+00:00'))
                diferenca = exp_date - pub_date
                
                print(f"   Diferen√ßa em dias: {diferenca.days}")
                if diferenca.days == 30:
                    print("   ‚úÖ Expira√ß√£o configurada corretamente para 30 dias")
                else:
                    print(f"   ‚ùå Erro: Expira√ß√£o deveria ser 30 dias, mas √© {diferenca.days}")
        
        # Testar filtro somente publicadas
        print("\n3. Testando filtro de empreendimentos publicados...")
        response = client.get('/empreendimentos?somente_publicadas=1')
        print(f"   Status: {response.status_code}")
        
        if response.status_code == 200:
            empreendimentos_publicados = response.get_json()
            print(f"   Empreendimentos publicados e n√£o expirados: {len(empreendimentos_publicados)}")
            
            # Verificar se o empreendimento rec√©m-publicado est√° na lista
            encontrado = any(emp.get('id') == empreendimento_id for emp in empreendimentos_publicados)
            if encontrado:
                print("   ‚úÖ Empreendimento publicado aparece na listagem")
            else:
                print("   ‚ùå Empreendimento publicado n√£o aparece na listagem")

def criar_planilha_grande(num_empreendimentos=50, unidades_por_emp=10):
    """Cria planilha grande para teste de performance"""
    wb = Workbook()
    ws = wb.active
    
    # Cabe√ßalhos
    headers = ['nome_empreendimento', 'nome_empresa', 'cep', 'endereco', 'observacao', 
               'numero_unidade', 'tamanho_m2', 'preco_venda', 'mecanismo_pagamento']
    
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    # Gerar dados
    mecanismos = ['financiamento', '√† vista', 'cons√≥rcio', 'outros']
    row_num = 2
    
    for emp_num in range(1, num_empreendimentos + 1):
        nome_emp = f'Residencial Performance {emp_num:03d}'
        nome_empresa = f'Construtora Performance {emp_num:03d} Ltda'
        cep = f'70{emp_num:03d}-000'
        endereco = f'Rua Performance {emp_num}, {emp_num * 100}'
        observacao = f'Empreendimento de teste de performance n√∫mero {emp_num}'
        
        for unidade_num in range(1, unidades_por_emp + 1):
            numero_unidade = f'{emp_num:03d}{unidade_num:02d}'
            tamanho_m2 = 45.0 + (unidade_num * 5)
            preco_venda = 120000 + (emp_num * 1000) + (unidade_num * 500)
            mecanismo = mecanismos[(emp_num + unidade_num) % len(mecanismos)]
            
            dados_linha = [nome_emp, nome_empresa, cep, endereco, observacao,
                          numero_unidade, tamanho_m2, preco_venda, mecanismo]
            
            for col_num, valor in enumerate(dados_linha, 1):
                ws.cell(row=row_num, column=col_num, value=valor)
            
            row_num += 1
    
    # Salvar em mem√≥ria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_performance():
    """Testa performance com planilha grande"""
    app = create_app()
    
    with app.test_client() as client:
        print("\nüöÄ Testando performance com planilha grande...")
        
        # Criar planilha com 20 empreendimentos e 5 unidades cada (100 linhas)
        print("\n1. Criando planilha de teste (20 empreendimentos, 5 unidades cada)...")
        planilha_grande = criar_planilha_grande(20, 5)
        
        # Teste de preview
        print("\n2. Testando preview de planilha grande...")
        inicio = datetime.now()
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_grande, 'performance_test.xlsx')},
                             content_type='multipart/form-data')
        
        fim = datetime.now()
        tempo_preview = (fim - inicio).total_seconds()
        
        print(f"   Status: {response.status_code}")
        print(f"   Tempo de preview: {tempo_preview:.2f} segundos")
        
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   Erros: {preview_data.get('erros')}")
            
            if tempo_preview < 5.0:
                print("   ‚úÖ Performance de preview aceit√°vel (< 5s)")
            else:
                print("   ‚ö†Ô∏è Performance de preview pode ser melhorada")
        
        # Teste de upload real
        print("\n3. Testando upload real de planilha grande...")
        planilha_grande_2 = criar_planilha_grande(20, 5)
        
        inicio = datetime.now()
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (planilha_grande_2, 'performance_test.xlsx')},
                             content_type='multipart/form-data')
        
        fim = datetime.now()
        tempo_upload = (fim - inicio).total_seconds()
        
        print(f"   Status: {response.status_code}")
        print(f"   Tempo de upload: {tempo_upload:.2f} segundos")
        
        if response.status_code == 200:
            upload_data = response.get_json()
            print(f"   Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
            print(f"   Unidades processadas: {upload_data.get('unidades_processadas')}")
            print(f"   Erros: {upload_data.get('erros')}")
            
            if tempo_upload < 10.0:
                print("   ‚úÖ Performance de upload aceit√°vel (< 10s)")
            else:
                print("   ‚ö†Ô∏è Performance de upload pode ser melhorada")

def test_edge_cases():
    """Testa casos extremos"""
    app = create_app()
    
    with app.test_client() as client:
        print("\nüîç Testando casos extremos...")
        
        # Teste 1: CEP com formata√ß√£o diferente
        print("\n1. Testando CEPs com formata√ß√µes diferentes...")
        ceps_teste = ['70000-000', '70000000', '70.000-000', '70 000 000']
        
        for i, cep in enumerate(ceps_teste, 1):
            empreendimento_teste = {
                'nome': f'Residencial CEP Teste {i}',
                'nome_empresa': f'Empresa CEP {i}',
                'cep': cep,
                'endereco': f'Rua CEP {i}, {i}00'
            }
            
            response = client.post('/empreendimentos',
                                 data=json.dumps(empreendimento_teste),
                                 content_type='application/json')
            
            print(f"   CEP '{cep}': Status {response.status_code}")
        
        # Teste 2: Valores monet√°rios com formata√ß√µes diferentes
        print("\n2. Testando valores monet√°rios...")
        valores_teste = [150000.00, 150000, '150000', '150.000,00', '150,000.00']
        
        # Criar empreendimento para as unidades
        empreendimento_valores = {
            'nome': 'Residencial Valores Teste',
            'nome_empresa': 'Empresa Valores',
            'cep': '70000-555',
            'endereco': 'Rua Valores, 555'
        }
        
        response = client.post('/empreendimentos',
                             data=json.dumps(empreendimento_valores),
                             content_type='application/json')
        
        if response.status_code == 201:
            empreendimento_id = response.get_json().get('id')
            
            for i, valor in enumerate(valores_teste, 1):
                unidade_teste = {
                    'empreendimento_id': empreendimento_id,
                    'numero_unidade': f'VAL{i}',
                    'tamanho_m2': 50.0,
                    'preco_venda': valor,
                    'mecanismo_pagamento': 'financiamento'
                }
                
                response = client.post('/api/unidades',
                                     data=json.dumps(unidade_teste),
                                     content_type='application/json')
                
                print(f"   Valor '{valor}': Status {response.status_code}")

def run_advanced_tests():
    """Executa testes avan√ßados"""
    print("üî¨ Iniciando testes avan√ßados do Sistema Obras HIS...")
    
    try:
        # Teste de expira√ß√£o
        test_expiracao_automatica()
        
        # Teste de performance
        test_performance()
        
        # Teste de casos extremos
        test_edge_cases()
        
        print("\n‚úÖ Todos os testes avan√ßados foram executados!")
        
    except Exception as e:
        print(f"\n‚ùå Erro durante os testes avan√ßados: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    run_advanced_tests()
