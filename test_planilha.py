import os
import sys
import io
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_obras_his.db'

# Importar e testar a aplicação
from app import create_app

def criar_planilha_teste():
    """Cria uma planilha de teste em memória"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos
    ws['A1'] = 'nome_empreendimento'
    ws['B1'] = 'nome_empresa'
    ws['C1'] = 'cep'
    ws['D1'] = 'endereco'
    ws['E1'] = 'observacao'
    ws['F1'] = 'numero_unidade'
    ws['G1'] = 'tamanho_m2'
    ws['H1'] = 'preco_venda'
    ws['I1'] = 'mecanismo_pagamento'
    
    # Dados de teste
    dados_teste = [
        ['Residencial Teste 1', 'Construtora ABC', '70000-001', 'Rua A, 123', 'Teste 1', '101', '50.5', '150000', 'financiamento'],
        ['Residencial Teste 1', 'Construtora ABC', '70000-001', 'Rua A, 123', 'Teste 1', '102', '60.0', '180000', 'à vista'],
        ['Residencial Teste 2', 'Construtora XYZ', '70000-002', 'Rua B, 456', 'Teste 2', '201', '45.0', '120000', 'consórcio'],
    ]
    
    for i, linha in enumerate(dados_teste, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em memória
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_planilha_functionality():
    """Testa funcionalidades de planilha"""
    app = create_app()
    
    with app.test_client() as client:
        print("🔍 Testando funcionalidades de planilha...")
        
        # Criar planilha de teste
        planilha = criar_planilha_teste()
        
        # Teste 1: Preview da planilha
        print("\n1. Testando preview da planilha...")
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha, 'teste.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   Preview válido: {preview_data.get('valido')}")
            print(f"   Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("   Detalhes dos erros:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"     - {erro}")
        else:
            print(f"   Erro: {response.get_json()}")
        
        # Teste 2: Upload real da planilha (se preview foi bem-sucedido)
        if response.status_code == 200:
            print("\n2. Testando upload real da planilha...")
            # Criar nova planilha para o upload real
            planilha_upload = criar_planilha_teste()
            
            response = client.post('/empreendimentos/upload',
                                 data={'file': (planilha_upload, 'teste.xlsx')},
                                 content_type='multipart/form-data')
            
            print(f"   Status: {response.status_code}")
            if response.status_code == 200:
                upload_data = response.get_json()
                print(f"   Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
                print(f"   Unidades processadas: {upload_data.get('unidades_processadas')}")
                print(f"   Erros: {upload_data.get('erros')}")
            else:
                print(f"   Erro: {response.get_json()}")
        
        # Teste 3: Verificar se os dados foram salvos
        print("\n3. Verificando dados salvos...")
        response = client.get('/empreendimentos')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   Total de empreendimentos: {len(empreendimentos)}")
            for emp in empreendimentos:
                print(f"     - {emp.get('nome')} ({emp.get('nome_empresa')})")
        
        print("\n✅ Testes de planilha concluídos!")

if __name__ == '__main__':
    test_planilha_functionality()
