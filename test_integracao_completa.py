import os
import sys
import requests
import json
import io
from openpyxl import Workbook
import time

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_integracao_obras_his.db'

def criar_planilha_teste():
    """Cria planilha de teste com dados variados"""
    wb = Workbook()
    ws = wb.active
    
    # CabeÃ§alhos
    ws['A1'] = 'Construtora'
    ws['B1'] = 'Empreendimento'
    ws['C1'] = 'EndereÃ§o'
    ws['D1'] = 'Observacao'
    ws['E1'] = 'NumeroUnidade'
    ws['F1'] = 'TamanhoM2'
    ws['G1'] = 'PrecoVenda'
    ws['H1'] = 'MecanismoPagamento'
    
    # Dados de teste
    dados = [
        ['DONATELO', 'TESTE', 'QR - 104 conjuto 4 - cep 72302004', 'Teste de integraÃ§Ã£o', '101', '50.5', '150000', 'financiamento'],
        ['CONSTRUTORA ABC', 'RESIDENCIAL FLORES', 'Rua das Flores, 123 - CEP: 70000-000', 'Empreendimento popular', '201', '45.0', '120000', 'Ã  vista'],
        ['EMPREENDIMENTOS XYZ', 'VILA ESPERANÃ‡A', 'Av. Principal, 456 - 71000111', 'HIS Categoria 1', '301', '55.2', '180000', 'consÃ³rcio'],
        ['DONATELO', 'TESTE', 'QR - 104 conjuto 4 - cep 72302004', '', '102', '48.0', '145000', 'outros'],
    ]
    
    for i, linha in enumerate(dados, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em memÃ³ria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_integracao_completa():
    """Teste de integraÃ§Ã£o completa do sistema"""
    print("ğŸš€ Iniciando teste de integraÃ§Ã£o completa...")
    
    base_url = "http://127.0.0.1:5000"
    
    # Teste 1: Health Check
    print("\n1. ğŸ” Testando Health Check...")
    try:
        response = requests.get(f"{base_url}/health")
        assert response.status_code == 200
        print("   âœ… Health check OK")
    except Exception as e:
        print(f"   âŒ Erro no health check: {e}")
        return False
    
    # Teste 2: Listar construtoras
    print("\n2. ğŸ¢ Testando listagem de construtoras...")
    try:
        response = requests.get(f"{base_url}/api/construtoras")
        assert response.status_code == 200
        construtoras = response.json()
        print(f"   âœ… {len(construtoras)} construtoras encontradas")
    except Exception as e:
        print(f"   âŒ Erro ao listar construtoras: {e}")
        return False
    
    # Teste 3: Preview da planilha
    print("\n3. ğŸ“Š Testando preview da planilha...")
    try:
        planilha = criar_planilha_teste()
        files = {'file': ('teste.xlsx', planilha, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        response = requests.post(f"{base_url}/empreendimentos/upload/preview", files=files)
        assert response.status_code == 200
        
        preview_data = response.json()
        print(f"   âœ… Preview vÃ¡lido: {preview_data.get('valido')}")
        print(f"   âœ… Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
        print(f"   âœ… Unidades encontradas: {preview_data.get('unidades_encontradas')}")
        print(f"   âœ… Erros: {preview_data.get('erros')}")
        
        if preview_data.get('detalhes_erros'):
            print("   âš ï¸ Detalhes dos erros:")
            for erro in preview_data.get('detalhes_erros', []):
                print(f"     - {erro}")
        
        # Verificar dados extraÃ­dos
        empreendimentos = preview_data.get('empreendimentos', [])
        for emp in empreendimentos:
            print(f"   ğŸ“‹ Empreendimento: {emp.get('nome')} - Empresa: {emp.get('nome_empresa')} - CEP: {emp.get('cep')}")
            
    except Exception as e:
        print(f"   âŒ Erro no preview da planilha: {e}")
        return False
    
    # Teste 4: Upload e processamento da planilha
    print("\n4. ğŸ“¤ Testando upload e processamento da planilha...")
    try:
        planilha = criar_planilha_teste()
        files = {'file': ('teste.xlsx', planilha, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        response = requests.post(f"{base_url}/empreendimentos/upload", files=files)
        assert response.status_code == 200
        
        upload_data = response.json()
        print(f"   âœ… Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
        print(f"   âœ… Unidades processadas: {upload_data.get('unidades_processadas')}")
        print(f"   âœ… Erros: {upload_data.get('erros')}")
        
        if upload_data.get('detalhes_erros'):
            print("   âš ï¸ Detalhes dos erros:")
            for erro in upload_data.get('detalhes_erros', []):
                print(f"     - {erro}")
                
    except Exception as e:
        print(f"   âŒ Erro no upload da planilha: {e}")
        return False
    
    # Teste 5: Listar empreendimentos criados
    print("\n5. ğŸ“‹ Testando listagem de empreendimentos...")
    try:
        response = requests.get(f"{base_url}/empreendimentos")
        assert response.status_code == 200
        
        empreendimentos = response.json()
        print(f"   âœ… {len(empreendimentos)} empreendimentos encontrados")
        
        for emp in empreendimentos:
            print(f"   ğŸ“ {emp.get('nome')} - {emp.get('nome_empresa')} - Status: {emp.get('status_publicacao', 'rascunho')}")
            
    except Exception as e:
        print(f"   âŒ Erro ao listar empreendimentos: {e}")
        return False
    
    # Teste 6: Publicar empreendimento
    print("\n6. ğŸ“¢ Testando publicaÃ§Ã£o de empreendimento...")
    try:
        if empreendimentos:
            emp_id = empreendimentos[0]['id']
            response = requests.post(f"{base_url}/empreendimentos/{emp_id}/publicar")
            assert response.status_code == 200
            
            pub_data = response.json()
            print(f"   âœ… Empreendimento {emp_id} publicado")
            print(f"   âœ… Publicado em: {pub_data.get('publicado_em')}")
            print(f"   âœ… Expira em: {pub_data.get('expira_em')}")
            
    except Exception as e:
        print(f"   âŒ Erro ao publicar empreendimento: {e}")
        return False
    
    # Teste 7: Filtrar apenas publicados
    print("\n7. ğŸ” Testando filtro de empreendimentos publicados...")
    try:
        response = requests.get(f"{base_url}/empreendimentos?somente_publicadas=1")
        assert response.status_code == 200
        
        publicados = response.json()
        print(f"   âœ… {len(publicados)} empreendimentos publicados encontrados")
        
    except Exception as e:
        print(f"   âŒ Erro ao filtrar publicados: {e}")
        return False
    
    # Teste 8: Listar unidades
    print("\n8. ğŸ  Testando listagem de unidades...")
    try:
        response = requests.get(f"{base_url}/api/unidades")
        assert response.status_code == 200
        
        unidades = response.json()
        print(f"   âœ… {len(unidades)} unidades encontradas")
        
        for unidade in unidades[:3]:  # Mostrar apenas as 3 primeiras
            print(f"   ğŸ  Unidade {unidade.get('numero_unidade')} - {unidade.get('tamanho_m2')}mÂ² - R$ {unidade.get('preco_venda')} - {unidade.get('mecanismo_pagamento')}")
            
    except Exception as e:
        print(f"   âŒ Erro ao listar unidades: {e}")
        return False
    
    # Teste 9: Teste de Unicode
    print("\n9. ğŸŒ Testando processamento Unicode...")
    try:
        # Criar empreendimento com caracteres especiais
        dados_unicode = {
            "nome": "Residencial SÃ£o JosÃ©",
            "nome_empresa": "ConstruÃ§Ã£o & Cia Ltda",
            "cep": "70000-000",
            "endereco": "Rua da ConceiÃ§Ã£o, 123 - Setor Habitacional",
            "observacao": "Empreendimento com acentuaÃ§Ã£o e sÃ­mbolos especiais: Ã§Ã£o, Ã£, Ã©, Ã¼"
        }
        
        response = requests.post(f"{base_url}/empreendimentos", json=dados_unicode)
        assert response.status_code == 201
        
        emp_criado = response.json()
        print(f"   âœ… Empreendimento Unicode criado: {emp_criado.get('nome')}")
        print(f"   âœ… Empresa: {emp_criado.get('nome_empresa')}")
        print(f"   âœ… ObservaÃ§Ã£o: {emp_criado.get('observacao')}")
        
    except Exception as e:
        print(f"   âŒ Erro no teste Unicode: {e}")
        return False
    
    print("\nğŸ‰ TODOS OS TESTES DE INTEGRAÃ‡ÃƒO PASSARAM!")
    print("\nğŸ“Š RESUMO DOS TESTES:")
    print("âœ… Health Check")
    print("âœ… Listagem de construtoras")
    print("âœ… Preview de planilha com extraÃ§Ã£o inteligente de CEP")
    print("âœ… Upload e processamento de planilha")
    print("âœ… Listagem de empreendimentos")
    print("âœ… Sistema de publicaÃ§Ã£o com expiraÃ§Ã£o")
    print("âœ… Filtros de empreendimentos")
    print("âœ… Listagem de unidades")
    print("âœ… Processamento Unicode")
    
    return True

if __name__ == '__main__':
    success = test_integracao_completa()
    if success:
        print("\nğŸš€ SISTEMA PRONTO PARA PRODUÃ‡ÃƒO!")
    else:
        print("\nâŒ Alguns testes falharam")
        sys.exit(1)
