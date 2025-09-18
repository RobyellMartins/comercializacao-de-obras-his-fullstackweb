import os
import sys
import io
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_usuario_obras_his.db'

# Importar e testar a aplica칞칚o
from app import create_app

def criar_planilha_usuario():
    """Cria planilha com os dados do usu치rio"""
    wb = Workbook()
    ws = wb.active
    
    # Cabe칞alhos conforme o usu치rio testou
    ws['A1'] = 'Construtora'
    ws['B1'] = 'Empreendimento'
    ws['C1'] = 'Endere칞o'
    
    # Dados do usu치rio
    ws['A2'] = 'DONATELO'
    ws['B2'] = 'TESTE'
    ws['C2'] = 'QR - 104 conjuto 4 - cep 72302004'
    
    # Salvar em mem칩ria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_planilha_usuario():
    """Testa planilha do usu치rio"""
    app = create_app()
    
    with app.test_client() as client:
        print("游댌 Testando planilha do usu치rio...")
        
        # Criar planilha do usu치rio
        planilha = criar_planilha_usuario()
        
        # Teste 1: Preview da planilha
        print("\n1. Testando preview da planilha do usu치rio...")
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha, 'teste_usuario.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   Preview v치lido: {preview_data.get('valido')}")
            print(f"   Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("   Detalhes dos erros:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"     - {erro}")
            
            # Mostrar dados encontrados
            empreendimentos = preview_data.get('empreendimentos', [])
            print(f"   Dados dos empreendimentos:")
            for emp in empreendimentos:
                print(f"     - Nome: {emp.get('nome')}")
                print(f"     - Empresa: {emp.get('nome_empresa')}")
                print(f"     - CEP: {emp.get('cep')}")
                print(f"     - Endere칞o: {emp.get('endereco')}")
        else:
            print(f"   Erro: {response.get_json()}")

if __name__ == '__main__':
    test_planilha_usuario()
