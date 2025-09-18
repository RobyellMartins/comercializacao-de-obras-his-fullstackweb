import os
import sys
import io
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_unicode_flexivel_obras_his.db'

# Importar e testar a aplica√ß√£o
from app import create_app

def criar_planilha_unicode_variada():
    """Cria planilha com caracteres unicode e diferentes formatos de colunas"""
    wb = Workbook()
    ws = wb.active
    
    # Cabe√ßalhos com diferentes varia√ß√µes e unicode
    ws['A1'] = 'Construtora'  # Varia√ß√£o 1
    ws['B1'] = 'Nome do Empreendimento'  # Varia√ß√£o 2
    ws['C1'] = 'Endere√ßo Completo'  # Varia√ß√£o 3
    ws['D1'] = 'Observa√ß√µes'  # Varia√ß√£o 4
    
    # Dados com caracteres unicode complexos
    dados_unicode = [
        [
            'Constru√ß√£o & Cia Ltda', 
            'Residencial S√£o Jos√© da Concei√ß√£o', 
            'Rua da Integra√ß√£o, 123 - Setor Habitacional - CEP: 72345-678', 
            'Empreendimento com acentua√ß√£o: √ß√£o, √£, √©, √º, √±, √ß'
        ],
        [
            'Empreendimentos A√ß√£o & Constru√ß√£o S.A.', 
            'Condom√≠nio √Åguas Cristalinas', 
            'Av. Jos√© de Alencar, 456 - cep 71234567', 
            'Observa√ß√£o com s√≠mbolos: R$ 150.000,00 - 50m¬≤ - √Årea de lazer'
        ],
        [
            'Construtora Uni√£o & Progresso', 
            'Vila Esperan√ßa - Habita√ß√£o Social', 
            'QR 104 Conjunto 4 Casa 15 - Samambaia/DF - 72302004', 
            'Projeto HIS: habita√ß√£o de interesse social com √°rea verde'
        ],
        [
            'Incorporadora S√£o Paulo & Bras√≠lia', 
            'Edif√≠cio Tr√™s Cora√ß√µes', 
            'SHIS QI 15 Conjunto 3 - Lago Sul - 71635030', 
            'Apartamentos de 2 e 3 quartos - √Årea nobre'
        ]
    ]
    
    for i, linha in enumerate(dados_unicode, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em mem√≥ria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def criar_planilha_formato_diferente():
    """Cria planilha com formato completamente diferente"""
    wb = Workbook()
    ws = wb.active
    
    # Cabe√ßalhos em formato diferente
    ws['A1'] = 'EMPRESA'  # Mai√∫sculo
    ws['B1'] = 'projeto'  # Min√∫sculo
    ws['C1'] = 'LOCAL'    # Diferente
    ws['D1'] = 'notas'    # Min√∫sculo
    ws['E1'] = 'apt'      # Abreviado
    ws['F1'] = '√°rea'     # Com acento
    ws['G1'] = 'valor'    # Diferente
    ws['H1'] = 'forma_pagto'  # Abreviado
    
    # Dados com unicode e formatos variados
    dados_variados = [
        [
            'Construtora Tr√™s Irm√£os & Filhos', 
            'Residencial P√¥r do Sol', 
            'Rua das Ac√°cias, 789 - Taguatinga - CEP 72000-000', 
            'Projeto sustent√°vel com energia solar',
            '101A',
            '65,5',
            'R$ 180.000,00',
            'financiamento'
        ],
        [
            'Empreendimentos S√£o Sebasti√£o', 
            'Condom√≠nio Jardim das Am√©ricas', 
            'QNM 25 Conjunto A - Ceil√¢ndia - 72220250', 
            '√Årea de lazer completa: piscina, quadra, sal√£o de festas',
            '202B',
            '58.0',
            '165000',
            '√† vista'
        ]
    ]
    
    for i, linha in enumerate(dados_variados, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em mem√≥ria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_unicode_planilha_flexivel():
    """Testa processamento unicode e flexibilidade de formato"""
    app = create_app()
    
    with app.test_client() as client:
        print("üåê Testando processamento Unicode e flexibilidade de planilhas...")
        
        # Teste 1: Planilha com unicode complexo
        print("\n1. üìä Testando planilha com caracteres Unicode complexos...")
        planilha_unicode = criar_planilha_unicode_variada()
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_unicode, 'unicode_test.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   ‚úÖ Preview v√°lido: {preview_data.get('valido')}")
            print(f"   ‚úÖ Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   ‚úÖ Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("   ‚ö†Ô∏è Detalhes dos erros:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"     - {erro}")
            
            # Verificar processamento unicode
            empreendimentos = preview_data.get('empreendimentos', [])
            print("\n   üìã Dados processados com Unicode:")
            for emp in empreendimentos:
                print(f"     üè¢ Nome: {emp.get('nome')}")
                print(f"     üèóÔ∏è Empresa: {emp.get('nome_empresa')}")
                print(f"     üìç CEP: {emp.get('cep')}")
                print(f"     üè† Endere√ßo: {emp.get('endereco')}")
                print(f"     üìù Observa√ß√£o: {emp.get('observacao')}")
                print("     ---")
        else:
            print(f"   ‚ùå Erro: {response.get_json()}")
            return False
        
        # Teste 2: Planilha com formato completamente diferente
        print("\n2. üîÑ Testando planilha com formato diferente...")
        planilha_diferente = criar_planilha_formato_diferente()
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_diferente, 'formato_diferente.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   ‚úÖ Preview v√°lido: {preview_data.get('valido')}")
            print(f"   ‚úÖ Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   ‚úÖ Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   ‚úÖ Erros: {preview_data.get('erros')}")
            
            # Verificar dados processados
            empreendimentos = preview_data.get('empreendimentos', [])
            unidades = preview_data.get('unidades', [])
            
            print("\n   üìã Empreendimentos processados:")
            for emp in empreendimentos:
                print(f"     üè¢ {emp.get('nome')} - {emp.get('nome_empresa')}")
                print(f"     üìç CEP: {emp.get('cep')} - Endere√ßo: {emp.get('endereco')}")
            
            print("\n   üè† Unidades processadas:")
            for unidade in unidades:
                print(f"     üè† Unidade {unidade.get('numero_unidade')} - {unidade.get('tamanho_m2')}m¬≤ - R$ {unidade.get('preco_venda')} - {unidade.get('mecanismo_pagamento')}")
        else:
            print(f"   ‚ùå Erro: {response.get_json()}")
            return False
        
        # Teste 3: Upload real com processamento completo
        print("\n3. üì§ Testando upload completo com processamento Unicode...")
        planilha_unicode = criar_planilha_unicode_variada()
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (planilha_unicode, 'upload_unicode.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            upload_data = response.get_json()
            print(f"   ‚úÖ Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
            print(f"   ‚úÖ Unidades processadas: {upload_data.get('unidades_processadas')}")
            print(f"   ‚úÖ Erros: {upload_data.get('erros')}")
            
            # Verificar se dados foram salvos corretamente
            response_list = client.get('/empreendimentos')
            if response_list.status_code == 200:
                empreendimentos_salvos = response_list.get_json()
                print(f"   ‚úÖ Total de empreendimentos no banco: {len(empreendimentos_salvos)}")
                
                # Verificar unicode nos dados salvos
                for emp in empreendimentos_salvos[-4:]:  # √öltimos 4 (os que acabamos de criar)
                    if emp.get('nome') and 'S√£o Jos√©' in emp.get('nome', ''):
                        print(f"   üåê Unicode preservado: {emp.get('nome')} - {emp.get('nome_empresa')}")
                        print(f"   üåê Observa√ß√£o: {emp.get('observacao')}")
        else:
            print(f"   ‚ùå Erro no upload: {response.get_json()}")
            return False
        
        # Teste 4: Verificar normaliza√ß√£o Unicode
        print("\n4. üî§ Testando normaliza√ß√£o Unicode...")
        
        # Criar dados com diferentes formas de unicode (NFC vs NFD)
        dados_teste = {
            "nome": "Residencial S√£o Jos√©",  # NFC
            "nome_empresa": "Constru√ß√£o & Cia Ltda",
            "cep": "70000-000",
            "endereco": "Rua da Concei√ß√£o, 123",
            "observacao": "Acentua√ß√£o: √ß√£o, √£, √©, √º, √±"
        }
        
        response = client.post('/empreendimentos', json=dados_teste)
        if response.status_code == 201:
            emp_criado = response.get_json()
            print(f"   ‚úÖ Empreendimento Unicode criado: {emp_criado.get('nome')}")
            print(f"   ‚úÖ Observa√ß√£o normalizada: {emp_criado.get('observacao')}")
        else:
            print(f"   ‚ùå Erro ao criar empreendimento Unicode: {response.get_json()}")
            return False
        
        print("\nüéâ TODOS OS TESTES DE UNICODE E FLEXIBILIDADE PASSARAM!")
        print("\nüìä RESUMO DOS TESTES:")
        print("‚úÖ Processamento de caracteres Unicode complexos")
        print("‚úÖ Extra√ß√£o de CEP de diferentes formatos")
        print("‚úÖ Mapeamento flex√≠vel de colunas")
        print("‚úÖ Normaliza√ß√£o Unicode (NFC)")
        print("‚úÖ Preserva√ß√£o de acentos e s√≠mbolos")
        print("‚úÖ Funcionamento independente do modelo da planilha")
        print("‚úÖ Upload e salvamento com Unicode")
        
        return True

if __name__ == '__main__':
    success = test_unicode_planilha_flexivel()
    if success:
        print("\nüåê SISTEMA VALIDADO: Unicode e flexibilidade de planilha funcionando perfeitamente!")
    else:
        print("\n‚ùå Alguns testes falharam")
        sys.exit(1)
