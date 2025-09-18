import os
import sys
from openpyxl import Workbook
import io

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_upload_debug.db'

# Importar e testar a aplicação
from app import create_app

def test_upload_planilha_debug():
    """Testa especificamente o upload de planilha para diagnosticar o problema"""
    app = create_app()
    
    with app.test_client() as client:
        print("🧪 TESTE DEBUG - UPLOAD DE PLANILHA")
        print("=" * 50)
        
        # Teste 1: Verificar se o backend está funcionando
        print("\n1. 🔧 Testando Backend...")
        response = client.get('/health')
        print(f"   ✅ Health Check: {response.status_code}")
        
        # Teste 2: Criar planilha de teste com formato correto
        print("\n2. 📊 Criando Planilha de Teste...")
        
        wb = Workbook()
        ws = wb.active
        
        # Cabeçalhos conforme esperado pelo sistema
        ws['A1'] = 'nome_empreendimento'
        ws['B1'] = 'nome_empresa'
        ws['C1'] = 'cep'
        ws['D1'] = 'endereco'
        ws['E1'] = 'observacao'
        ws['F1'] = 'numero_unidade'
        ws['G1'] = 'tamanho_m2'
        ws['H1'] = 'preco_venda'
        ws['I1'] = 'mecanismo_pagamento'
        
        # Dados de teste
        ws['A2'] = 'Residencial Teste Upload'
        ws['B2'] = 'Construtora Upload Ltda'
        ws['C2'] = '70000-000'
        ws['D2'] = 'Rua do Upload, 123 - Brasília/DF'
        ws['E2'] = 'Empreendimento para teste de upload'
        ws['F2'] = '101'
        ws['G2'] = 65.5
        ws['H2'] = 180000.00
        ws['I2'] = 'financiamento'
        
        # Segunda linha - mesma empresa, unidade diferente
        ws['A3'] = 'Residencial Teste Upload'
        ws['B3'] = 'Construtora Upload Ltda'
        ws['C3'] = '70000-000'
        ws['D3'] = 'Rua do Upload, 123 - Brasília/DF'
        ws['E3'] = 'Empreendimento para teste de upload'
        ws['F3'] = '102'
        ws['G3'] = 58.0
        ws['H3'] = 165000.00
        ws['I3'] = 'à vista'
        
        # Terceira linha - empreendimento diferente
        ws['A4'] = 'Condomínio Teste Upload'
        ws['B4'] = 'Outra Construtora Ltda'
        ws['C4'] = '71000-000'
        ws['D4'] = 'Av. do Teste, 456 - Brasília/DF'
        ws['E4'] = 'Segundo empreendimento de teste'
        ws['F4'] = '201'
        ws['G4'] = 72.0
        ws['H4'] = 200000.00
        ws['I4'] = 'consórcio'
        
        print("   📝 Planilha criada com:")
        print("   - 2 empreendimentos")
        print("   - 3 unidades")
        print("   - Cabeçalhos corretos")
        
        # Salvar em memória
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        # Teste 3: Testar preview primeiro
        print("\n3. 👁️ Testando Preview da Planilha...")
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (file_stream, 'teste_upload.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   ✅ Preview Status: {response.status_code}")
        
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   📊 Preview válido: {preview_data.get('valido')}")
            print(f"   📊 Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   📊 Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   📊 Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("   ❌ Erros encontrados:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"      - {erro}")
            
            if preview_data.get('empreendimentos'):
                print("   📋 Empreendimentos no preview:")
                for emp in preview_data.get('empreendimentos', []):
                    print(f"      - {emp.get('nome')} - {emp.get('nome_empresa')}")
        else:
            print(f"   ❌ Erro no preview: {response.get_json()}")
        
        # Teste 4: Testar upload real
        print("\n4. 📤 Testando Upload Real...")
        
        # Resetar stream
        file_stream.seek(0)
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (file_stream, 'teste_upload.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   ✅ Upload Status: {response.status_code}")
        
        if response.status_code == 200:
            upload_data = response.get_json()
            print(f"   📊 Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
            print(f"   📊 Unidades processadas: {upload_data.get('unidades_processadas')}")
            print(f"   📊 Erros: {upload_data.get('erros')}")
            
            if upload_data.get('detalhes_erros'):
                print("   ❌ Erros encontrados:")
                for erro in upload_data.get('detalhes_erros', []):
                    print(f"      - {erro}")
            
            if upload_data.get('empreendimentos'):
                print("   📋 Empreendimentos criados:")
                for emp in upload_data.get('empreendimentos', []):
                    print(f"      - ID: {emp.get('id')} - {emp.get('nome')}")
            
            if upload_data.get('unidades'):
                print("   🏠 Unidades criadas:")
                for unidade in upload_data.get('unidades', []):
                    print(f"      - Unidade {unidade.get('numero_unidade')} - Emp ID: {unidade.get('empreendimento_id')}")
        else:
            error_data = response.get_json()
            print(f"   ❌ Erro no upload: {error_data}")
        
        # Teste 5: Verificar se dados foram salvos no banco
        print("\n5. 🗄️ Verificando Dados no Banco...")
        
        response = client.get('/empreendimentos')
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   📊 Total de empreendimentos no banco: {len(empreendimentos)}")
            
            for emp in empreendimentos:
                print(f"      - {emp.get('nome')} - {emp.get('nome_empresa')}")
        
        response = client.get('/api/unidades')
        if response.status_code == 200:
            unidades = response.get_json()
            print(f"   📊 Total de unidades no banco: {len(unidades)}")
            
            for unidade in unidades:
                print(f"      - Unidade {unidade.get('numero_unidade')} - {unidade.get('mecanismo_pagamento')}")
        
        # Teste 6: Testar com planilha com cabeçalhos alternativos
        print("\n6. 🔄 Testando com Cabeçalhos Alternativos...")
        
        wb2 = Workbook()
        ws2 = wb2.active
        
        # Cabeçalhos alternativos que deveriam funcionar
        ws2['A1'] = 'empreendimento'  # alternativo para nome_empreendimento
        ws2['B1'] = 'construtora'     # alternativo para nome_empresa
        ws2['C1'] = 'cep'
        ws2['D1'] = 'endereço'        # com acento
        ws2['E1'] = 'obs'             # alternativo para observacao
        ws2['F1'] = 'unidade'         # alternativo para numero_unidade
        ws2['G1'] = 'area'            # alternativo para tamanho_m2
        ws2['H1'] = 'preco'           # alternativo para preco_venda
        ws2['I1'] = 'pagamento'       # alternativo para mecanismo_pagamento
        
        # Dados
        ws2['A2'] = 'Residencial Cabeçalhos Alt'
        ws2['B2'] = 'Construtora Alt Ltda'
        ws2['C2'] = '72000-000'
        ws2['D2'] = 'Rua Alternativa, 789'
        ws2['E2'] = 'Teste com cabeçalhos alternativos'
        ws2['F2'] = '301'
        ws2['G2'] = 80.0
        ws2['H2'] = 220000.00
        ws2['I2'] = 'outros'
        
        file_stream2 = io.BytesIO()
        wb2.save(file_stream2)
        file_stream2.seek(0)
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (file_stream2, 'teste_alt.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   ✅ Preview Alternativo Status: {response.status_code}")
        
        if response.status_code == 200:
            preview_alt = response.get_json()
            print(f"   📊 Preview alternativo válido: {preview_alt.get('valido')}")
            print(f"   📊 Empreendimentos encontrados: {preview_alt.get('empreendimentos_encontrados')}")
            
            if not preview_alt.get('valido'):
                print("   ❌ Erros com cabeçalhos alternativos:")
                for erro in preview_alt.get('detalhes_erros', []):
                    print(f"      - {erro}")
        
        print("\n" + "=" * 50)
        print("🎯 DIAGNÓSTICO COMPLETO!")
        
        return True

if __name__ == '__main__':
    try:
        success = test_upload_planilha_debug()
        if success:
            print("\n✅ Diagnóstico concluído!")
        else:
            print("\n❌ Erro no diagnóstico")
            sys.exit(1)
    except Exception as e:
        print(f"\n❌ Erro durante o diagnóstico: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
