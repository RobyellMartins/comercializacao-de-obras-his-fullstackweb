#!/usr/bin/env python3
"""
Diagn√≥stico completo do sistema de upload de planilhas
Testa todas as poss√≠veis causas do erro "Erro ao fazer upload do arquivo. Tente novamente."
"""

import sys
import os
import requests
import json
from io import BytesIO
import tempfile
from openpyxl import Workbook

# Adicionar diret√≥rio raiz ao path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def criar_planilha_teste():
    """Cria uma planilha de teste simples"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws['A1'] = 'NomeEmpreendimento'
    ws['B1'] = 'Endereco'
    ws['C1'] = 'NomeEmpresa'
    ws['D1'] = 'NumeroUnidade'
    ws['E1'] = 'TamanhoM2'

    # Dados de teste
    ws['A2'] = 'Teste Diagn√≥stico'
    ws['B2'] = 'Rua Teste, 123 - Centro - S√£o Paulo/SP - CEP: 01234-567'
    ws['C2'] = 'Construtora Teste'
    ws['D2'] = '101'
    ws['E2'] = '75.5'

    # Salvar em mem√≥ria
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def testar_conexao_servidor():
    """Testa se o servidor est√° respondendo"""
    print("üîç Testando conex√£o com o servidor...")

    try:
        # Testar endpoint de sa√∫de
        response = requests.get("http://localhost:5000/health", timeout=10)
        if response.status_code == 200:
            print("‚úÖ Servidor est√° respondendo")
            return True
        else:
            print(f"‚ùå Servidor respondeu com status {response.status_code}")
            return False
    except requests.exceptions.ConnectionError:
        print("‚ùå N√£o foi poss√≠vel conectar ao servidor (servidor pode n√£o estar rodando)")
        return False
    except Exception as e:
        print(f"‚ùå Erro ao conectar: {str(e)}")
        return False

def testar_endpoint_upload():
    """Testa o endpoint de upload diretamente"""
    print("\nüîç Testando endpoint de upload...")

    try:
        # Criar planilha de teste
        planilha_data = criar_planilha_teste()

        # Preparar dados do formul√°rio
        files = {'file': ('teste.xlsx', planilha_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        # Fazer upload
        response = requests.post(
            "http://localhost:5000/empreendimentos/upload",
            files=files,
            timeout=30
        )

        print(f"Status da resposta: {response.status_code}")

        if response.status_code == 200:
            print("‚úÖ Upload bem-sucedido!")
            resultado = response.json()
            print(f"   Empreendimentos processados: {resultado.get('empreendimentos_processados', 0)}")
            print(f"   Unidades processadas: {resultado.get('unidades_processadas', 0)}")
            return True
        else:
            print(f"‚ùå Upload falhou com status {response.status_code}")
            try:
                erro = response.json()
                print(f"   Erro: {erro.get('error', 'Erro desconhecido')}")
                if 'detalhes' in erro:
                    print(f"   Detalhes: {erro['detalhes']}")
            except:
                print(f"   Resposta: {response.text[:200]}...")
            return False

    except requests.exceptions.Timeout:
        print("‚ùå Timeout no upload (servidor pode estar sobrecarregado)")
        return False
    except Exception as e:
        print(f"‚ùå Erro no upload: {str(e)}")
        return False

def testar_cors():
    """Testa configura√ß√£o CORS"""
    print("\nüîç Testando configura√ß√£o CORS...")

    try:
        # Fazer uma requisi√ß√£o OPTIONS (preflight)
        response = requests.options(
            "http://localhost:5000/empreendimentos/upload",
            headers={
                'Origin': 'http://localhost:3000',
                'Access-Control-Request-Method': 'POST',
                'Access-Control-Request-Headers': 'Content-Type'
            },
            timeout=10
        )

        cors_headers = ['access-control-allow-origin', 'access-control-allow-methods', 'access-control-allow-headers']

        cors_present = any(h in response.headers for h in cors_headers)
        if cors_present:
            print("‚úÖ CORS configurado corretamente")
            return True
        else:
            print("‚ùå Headers CORS n√£o encontrados")
            print("   Headers de resposta:", list(response.headers.keys()))
            return False

    except Exception as e:
        print(f"‚ùå Erro ao testar CORS: {str(e)}")
        return False

def testar_tamanho_arquivo():
    """Testa limite de tamanho de arquivo"""
    print("\nüîç Testando limite de tamanho de arquivo...")

    try:
        # Criar arquivo grande (mas ainda dentro do limite)
        wb = Workbook()
        ws = wb.active

        # Adicionar muitos dados
        for i in range(1, 1000):
            ws[f'A{i}'] = f'Dado {i}'
            ws[f'B{i}'] = f'Endere√ßo {i}'
            ws[f'C{i}'] = f'Empresa {i}'

        bio = BytesIO()
        wb.save(bio)
        tamanho = len(bio.getvalue()) / 1024 / 1024  # MB
        print(f"   Tamanho do arquivo de teste: {tamanho:.2f} MB")

        bio.seek(0)
        files = {'file': ('teste_grande.xlsx', bio, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        response = requests.post(
            "http://localhost:5000/empreendimentos/upload",
            files=files,
            timeout=60
        )

        if response.status_code == 413:
            print("‚ùå Arquivo muito grande (413 - Payload Too Large)")
            return False
        elif response.status_code == 200:
            print("‚úÖ Arquivo dentro do limite de tamanho")
            return True
        else:
            print(f"‚ö†Ô∏è  Status inesperado: {response.status_code}")
            return True

    except Exception as e:
        print(f"‚ùå Erro ao testar tamanho: {str(e)}")
        return False

def testar_banco_dados():
    """Testa conex√£o com banco de dados"""
    print("\nüîç Testando conex√£o com banco de dados...")

    try:
        from app import create_app
        app = create_app()

        with app.app_context():
            from src.db_sql import get_db
            with get_db() as db:
                # Testar uma query simples
                result = db.execute("SELECT 1").fetchone()
                if result:
                    print("‚úÖ Conex√£o com banco de dados OK")
                    return True
                else:
                    print("‚ùå Query de teste falhou")
                    return False

    except Exception as e:
        print(f"‚ùå Erro na conex√£o com banco: {str(e)}")
        return False

def testar_logs():
    """Verifica se h√° logs de erro recentes"""
    print("\nüîç Verificando logs de erro...")

    try:
        log_files = ['logs/flask.log', 'logs/error.log']
        found_errors = False

        for log_file in log_files:
            if os.path.exists(log_file):
                print(f"   Verificando {log_file}...")
                with open(log_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()[-20:]  # √öltimas 20 linhas

                    for line in lines:
                        if 'ERROR' in line or 'Exception' in line:
                            print(f"   ‚ùå Erro encontrado: {line.strip()}")
                            found_errors = True

        if not found_errors:
            print("‚úÖ Nenhum erro recente encontrado nos logs")
            return True
        else:
            return False

    except Exception as e:
        print(f"‚ùå Erro ao verificar logs: {str(e)}")
        return False

def executar_diagnostico_completo():
    """Executa todos os testes de diagn√≥stico"""
    print("üöÄ INICIANDO DIAGN√ìSTICO COMPLETO DO SISTEMA DE UPLOAD")
    print("=" * 60)

    testes = [
        ("Conex√£o com Servidor", testar_conexao_servidor),
        ("Configura√ß√£o CORS", testar_cors),
        ("Conex√£o com Banco", testar_banco_dados),
        ("Limite de Tamanho", testar_tamanho_arquivo),
        ("Endpoint de Upload", testar_endpoint_upload),
        ("Logs de Erro", testar_logs),
    ]

    resultados = {}

    for nome_teste, funcao_teste in testes:
        try:
            resultado = funcao_teste()
            resultados[nome_teste] = resultado
        except Exception as e:
            print(f"‚ùå Erro ao executar teste '{nome_teste}': {str(e)}")
            resultados[nome_teste] = False

    print("\n" + "=" * 60)
    print("üìä RESUMO DOS TESTES:")

    todos_passaram = True
    for nome_teste, resultado in resultados.items():
        status = "‚úÖ PASSOU" if resultado else "‚ùå FALHOU"
        print(f"   {nome_teste}: {status}")
        if not resultado:
            todos_passaram = False

    print("\n" + "=" * 60)
    if todos_passaram:
        print("üéâ TODOS OS TESTES PASSARAM!")
        print("   Se ainda h√° erro de upload, pode ser:")
        print("   - Problema no frontend (JavaScript/React)")
        print("   - Arquivo do usu√°rio com formato diferente")
        print("   - Problema de rede entre frontend e backend")
    else:
        print("‚ùå ALGUNS TESTES FALHARAM!")
        print("   Os testes que falharam precisam ser corrigidos primeiro.")

    return todos_passaram

if __name__ == "__main__":
    sucesso = executar_diagnostico_completo()
    print(f"\nResultado final: {'PASSOU' if sucesso else 'FALHOU'}")
    sys.exit(0 if sucesso else 1)
