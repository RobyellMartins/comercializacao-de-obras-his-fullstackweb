import os
import sys
import io
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_unicode_flexivel_obras_his.db'

# Importar e testar a aplicação
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app

def criar_planilha_unicode_variada():
    """Cria planilha com caracteres unicode e diferentes formatos de colunas"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos com diferentes variações e unicode
    ws['A1'] = 'Construtora'  # Variação 1
    ws['B1'] = 'Nome do Empreendimento'  # Variação 2
    ws['C1'] = 'Endereço Completo'  # Variação 3
    ws['D1'] = 'Observações'  # Variação 4
    
    # Dados com caracteres unicode complexos
    dados_unicode = [
        [
            'Construção & Cia Ltda', 
            'Residencial São José da Conceição', 
            'Rua da Integração, 123 - Setor Habitacional - CEP: 72345-678', 
            'Empreendimento com acentuação: ção, ã, é, ü, ñ, ç'
        ],
        [
            'Empreendimentos Ação & Construção S.A.', 
            'Condomínio Águas Cristalinas', 
            'Av. José de Alencar, 456 - cep 71234567', 
            'Observação com símbolos: R$ 150.000,00 - 50m² - Área de lazer'
        ],
        [
            'Construtora União & Progresso', 
            'Vila Esperança - Habitação Social', 
            'QR 104 Conjunto 4 Casa 15 - Samambaia/DF - 72302004', 
            'Projeto HIS: habitação de interesse social com área verde'
        ],
        [
            'Incorporadora São Paulo & Brasília', 
            'Edifício Três Corações', 
            'SHIS QI 15 Conjunto 3 - Lago Sul - 71635030', 
            'Apartamentos de 2 e 3 quartos - Área nobre'
        ]
    ]
    
    for i, linha in enumerate(dados_unicode, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em memória
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def criar_planilha_formato_diferente():
    """Cria planilha com formato completamente diferente"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos em formato diferente
    ws['A1'] = 'EMPRESA'  # Maiúsculo
    ws['B1'] = 'projeto'  # Minúsculo
    ws['C1'] = 'LOCAL'    # Diferente
    ws['D1'] = 'notas'    # Minúsculo
    ws['E1'] = 'apt'      # Abreviado
    ws['F1'] = 'área'     # Com acento
    ws['G1'] = 'valor'    # Diferente
    ws['H1'] = 'forma_pagto'  # Abreviado
    
    # Dados com unicode e formatos variados
    dados_variados = [
        [
            'Construtora Três Irmãos & Filhos', 
            'Residencial Pôr do Sol', 
            'Rua das Acácias, 789 - Taguatinga - CEP 72000-000', 
            'Projeto sustentável com energia solar',
            '101A',
            '65,5',
            'R$ 180.000,00',
            'financiamento'
        ],
        [
            'Empreendimentos São Sebastião', 
            'Condomínio Jardim das Américas', 
            'QNM 25 Conjunto A - Ceilândia - 72220250', 
            'Área de lazer completa: piscina, quadra, salão de festas',
            '202B',
            '58.0',
            '165000',
            'à vista'
        ]
    ]
    
    for i, linha in enumerate(dados_variados, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em memória
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_unicode_planilha_flexivel():
    """Testa processamento unicode e flexibilidade de formato"""
    app = create_app()
    
    with app.test_client() as client:
        print(" Testando processamento Unicode e flexibilidade de planilhas...")
        
        # Teste 1: Planilha com unicode complexo
        print("\n1.  Testando planilha com caracteres Unicode complexos...")
        planilha_unicode = criar_planilha_unicode_variada()
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_unicode, 'unicode_test.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"    Preview válido: {preview_data.get('valido')}")
            print(f"    Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"    Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("    Detalhes dos erros:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"     - {erro}")
            
            # Verificar processamento unicode
            empreendimentos = preview_data.get('empreendimentos', [])
            print("\n    Dados processados com Unicode:")
            for emp in empreendimentos:
                print(f"      Nome: {emp.get('nome')}")
                print(f"      Empresa: {emp.get('nome_empresa')}")
                print(f"      CEP: {emp.get('cep')}")
                print(f"      Endereço: {emp.get('endereco')}")
                print(f"      Observação: {emp.get('observacao')}")
                print("     ---")
        else:
            print(f"    Erro: {response.get_json()}")
            return False
        
        # Teste 2: Planilha com formato completamente diferente
        print("\n2.  Testando planilha com formato diferente...")
        planilha_diferente = criar_planilha_formato_diferente()
        
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_diferente, 'formato_diferente.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"    Preview válido: {preview_data.get('valido')}")
            print(f"    Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"    Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"    Erros: {preview_data.get('erros')}")
            
            # Verificar dados processados
            empreendimentos = preview_data.get('empreendimentos', [])
            unidades = preview_data.get('unidades', [])
            
            print("\n    Empreendimentos processados:")
            for emp in empreendimentos:
                print(f"      {emp.get('nome')} - {emp.get('nome_empresa')}")
                print(f"      CEP: {emp.get('cep')} - Endereço: {emp.get('endereco')}")
            
            print("\n    Unidades processadas:")
            for unidade in unidades:
                print(f"      Unidade {unidade.get('numero_unidade')} - {unidade.get('tamanho_m2')}m² - R$ {unidade.get('preco_venda')} - {unidade.get('mecanismo_pagamento')}")
        else:
            print(f"    Erro: {response.get_json()}")
            return False
        
        # Teste 3: Upload real com processamento completo
        print("\n3.  Testando upload completo com processamento Unicode...")
        planilha_unicode = criar_planilha_unicode_variada()
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (planilha_unicode, 'upload_unicode.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            upload_data = response.get_json()
            print(f"    Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
            print(f"    Unidades processadas: {upload_data.get('unidades_processadas')}")
            print(f"    Erros: {upload_data.get('erros')}")
            
            # Verificar se dados foram salvos corretamente
            response_list = client.get('/empreendimentos')
            if response_list.status_code == 200:
                empreendimentos_salvos = response_list.get_json()
                print(f"    Total de empreendimentos no banco: {len(empreendimentos_salvos)}")
                
                # Verificar unicode nos dados salvos
                for emp in empreendimentos_salvos[-4:]:  # Últimos 4 (os que acabamos de criar)
                    if emp.get('nome') and 'São José' in emp.get('nome', ''):
                        print(f"    Unicode preservado: {emp.get('nome')} - {emp.get('nome_empresa')}")
                        print(f"    Observação: {emp.get('observacao')}")
        else:
            print(f"    Erro no upload: {response.get_json()}")
            return False
        
        # Teste 4: Verificar normalização Unicode
        print("\n4.  Testando normalização Unicode...")
        
        # Criar dados com diferentes formas de unicode (NFC vs NFD)
        dados_teste = {
            "nome": "Residencial São José",  # NFC
            "nome_empresa": "Construção & Cia Ltda",
            "cep": "70000-000",
            "endereco": "Rua da Conceição, 123",
            "observacao": "Acentuação: ção, ã, é, ü, ñ"
        }
        
        response = client.post('/empreendimentos', json=dados_teste)
        if response.status_code == 201:
            emp_criado = response.get_json()
            print(f"    Empreendimento Unicode criado: {emp_criado.get('nome')}")
            print(f"    Observação normalizada: {emp_criado.get('observacao')}")
        else:
            print(f"    Erro ao criar empreendimento Unicode: {response.get_json()}")
            return False
        
        print("\n TODOS OS TESTES DE UNICODE E FLEXIBILIDADE PASSARAM!")
        print("\n RESUMO DOS TESTES:")
        print(" Processamento de caracteres Unicode complexos")
        print(" Extração de CEP de diferentes formatos")
        print(" Mapeamento flexível de colunas")
        print(" Normalização Unicode (NFC)")
        print(" Preservação de acentos e símbolos")
        print(" Funcionamento independente do modelo da planilha")
        print(" Upload e salvamento com Unicode")
        
        return True

if __name__ == '__main__':
    success = test_unicode_planilha_flexivel()
    if success:
        print("\n SISTEMA VALIDADO: Unicode e flexibilidade de planilha funcionando perfeitamente!")
    else:
        print("\n Alguns testes falharam")
        sys.exit(1)
