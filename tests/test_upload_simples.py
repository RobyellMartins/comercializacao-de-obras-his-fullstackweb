import os
import sys
from openpyxl import Workbook
import io

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_upload_simples.db'

# Importar e testar a aplicação
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app

def test_upload_simples():
    """Teste simples e direto do upload de planilha"""
    app = create_app()
    
    with app.test_client() as client:
        print(" TESTE SIMPLES - UPLOAD DE PLANILHA")
        print("=" * 50)
        
        # Criar planilha de teste
        wb = Workbook()
        ws = wb.active
        
        # Cabeçalhos
        ws['A1'] = 'nome_empreendimento'
        ws['B1'] = 'nome_empresa'
        ws['C1'] = 'cep'
        ws['D1'] = 'endereco'
        ws['E1'] = 'observacao'
        ws['F1'] = 'numero_unidade'
        ws['G1'] = 'tamanho_m2'
        ws['H1'] = 'preco_venda'
        ws['I1'] = 'mecanismo_pagamento'
        
        # Dados
        ws['A2'] = 'Residencial Teste'
        ws['B2'] = 'Construtora Teste'
        ws['C2'] = '70000-000'
        ws['D2'] = 'Rua Teste, 123'
        ws['E2'] = 'Teste'
        ws['F2'] = '101'
        ws['G2'] = 65.5
        ws['H2'] = 180000.00
        ws['I2'] = 'financiamento'
        
        print(" Planilha criada com 1 empreendimento e 1 unidade")
        
        # Teste 1: Preview
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        print("\n1.  Testando Preview...")
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (file_stream, 'teste.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            data = response.get_json()
            print(f"   Válido: {data.get('valido')}")
            print(f"   Empreendimentos: {data.get('empreendimentos_encontrados')}")
            print(f"   Unidades: {data.get('unidades_encontradas')}")
        
        # Teste 2: Upload real - criar novo stream
        file_stream2 = io.BytesIO()
        wb.save(file_stream2)
        file_stream2.seek(0)
        
        print("\n2.  Testando Upload Real...")
        response = client.post('/empreendimentos/upload',
                             data={'file': (file_stream2, 'teste.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            data = response.get_json()
            print(f"   Empreendimentos processados: {data.get('empreendimentos_processados')}")
            print(f"   Unidades processadas: {data.get('unidades_processadas')}")
            print(f"   Erros: {data.get('erros')}")
            
            if data.get('detalhes_erros'):
                print("   Erros encontrados:")
                for erro in data.get('detalhes_erros'):
                    print(f"      - {erro}")
        else:
            print(f"   Erro: {response.get_json()}")
        
        # Teste 3: Verificar banco
        print("\n3.  Verificando Banco...")
        response = client.get('/empreendimentos')
        if response.status_code == 200:
            emps = response.get_json()
            print(f"   Empreendimentos no banco: {len(emps)}")
        
        response = client.get('/api/unidades')
        if response.status_code == 200:
            unis = response.get_json()
            print(f"   Unidades no banco: {len(unis)}")
        
        return True

if __name__ == '__main__':
    try:
        test_upload_simples()
    except Exception as e:
        print(f"Erro: {str(e)}")
        import traceback
        traceback.print_exc()
