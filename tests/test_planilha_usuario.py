import os
import sys
import io
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_usuario_obras_his.db'

# Importar e testar a aplicação
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app

def criar_planilha_usuario():
    """Cria planilha com os dados do usuário"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos conforme o usuário testou
    ws['A1'] = 'Construtora'
    ws['B1'] = 'Empreendimento'
    ws['C1'] = 'Endereço'
    
    # Dados do usuário
    ws['A2'] = 'DONATELO'
    ws['B2'] = 'TESTE'
    ws['C2'] = 'QR - 104 conjuto 4 - cep 72302004'
    
    # Salvar em memória
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_planilha_usuario():
    """Testa planilha do usuário"""
    app = create_app()
    
    with app.test_client() as client:
        print(" Testando planilha do usuário...")
        
        # Criar planilha do usuário
        planilha = criar_planilha_usuario()
        
        # Teste 1: Preview da planilha
        print("\n1. Testando preview da planilha do usuário...")
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha, 'teste_usuario.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   Preview válido: {preview_data.get('valido')}")
            print(f"   Empreendimentos encontrados: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   Unidades encontradas: {preview_data.get('unidades_encontradas')}")
            print(f"   Erros: {preview_data.get('erros')}")
            
            if preview_data.get('detalhes_erros'):
                print("   Detalhes dos erros:")
                for erro in preview_data.get('detalhes_erros', []):
                    print(f"     - {erro}")
            
            # Mostrar dados encontrados
            empreendimentos = preview_data.get('empreendimentos', [])
            print(f"   Dados dos empreendimentos:")
            for emp in empreendimentos:
                print(f"     - Nome: {emp.get('nome')}")
                print(f"     - Empresa: {emp.get('nome_empresa')}")
                print(f"     - CEP: {emp.get('cep')}")
                print(f"     - Endereço: {emp.get('endereco')}")
        else:
            print(f"   Erro: {response.get_json()}")

if __name__ == '__main__':
    test_planilha_usuario()
