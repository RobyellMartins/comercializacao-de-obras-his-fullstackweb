#!/usr/bin/env python3
"""
Diagnóstico completo do sistema de upload de planilhas
Testa todas as possíveis causas do erro "Erro ao fazer upload do arquivo. Tente novamente."
"""

import sys
import os
import requests
import json
from io import BytesIO
import tempfile
from openpyxl import Workbook

# Adicionar diretório raiz ao path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def criar_planilha_teste():
    """Cria uma planilha de teste simples"""
    wb = Workbook()
    ws = wb.active

    # Headers
    ws['A1'] = 'NomeEmpreendimento'
    ws['B1'] = 'Endereco'
    ws['C1'] = 'NomeEmpresa'
    ws['D1'] = 'NumeroUnidade'
    ws['E1'] = 'TamanhoM2'

    # Dados de teste
    ws['A2'] = 'Teste Diagnóstico'
    ws['B2'] = 'Rua Teste, 123 - Centro - São Paulo/SP - CEP: 01234-567'
    ws['C2'] = 'Construtora Teste'
    ws['D2'] = '101'
    ws['E2'] = '75.5'

    # Salvar em memória
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def testar_conexao_servidor():
    """Testa se o servidor está respondendo"""
    print(" Testando conexão com o servidor...")

    try:
        # Testar endpoint de saúde
        response = requests.get("http://localhost:5000/health", timeout=10)
        if response.status_code == 200:
            print(" Servidor está respondendo")
            return True
        else:
            print(f" Servidor respondeu com status {response.status_code}")
            return False
    except requests.exceptions.ConnectionError:
        print(" Não foi possível conectar ao servidor (servidor pode não estar rodando)")
        return False
    except Exception as e:
        print(f" Erro ao conectar: {str(e)}")
        return False

def testar_endpoint_upload():
    """Testa o endpoint de upload diretamente"""
    print("\n Testando endpoint de upload...")

    try:
        # Criar planilha de teste
        planilha_data = criar_planilha_teste()

        # Preparar dados do formulário
        files = {'file': ('teste.xlsx', planilha_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        # Fazer upload
        response = requests.post(
            "http://localhost:5000/empreendimentos/upload",
            files=files,
            timeout=30
        )

        print(f"Status da resposta: {response.status_code}")

        if response.status_code == 200:
            print(" Upload bem-sucedido!")
            resultado = response.json()
            print(f"   Empreendimentos processados: {resultado.get('empreendimentos_processados', 0)}")
            print(f"   Unidades processadas: {resultado.get('unidades_processadas', 0)}")
            return True
        else:
            print(f" Upload falhou com status {response.status_code}")
            try:
                erro = response.json()
                print(f"   Erro: {erro.get('error', 'Erro desconhecido')}")
                if 'detalhes' in erro:
                    print(f"   Detalhes: {erro['detalhes']}")
            except:
                print(f"   Resposta: {response.text[:200]}...")
            return False

    except requests.exceptions.Timeout:
        print(" Timeout no upload (servidor pode estar sobrecarregado)")
        return False
    except Exception as e:
        print(f" Erro no upload: {str(e)}")
        return False

def testar_cors():
    """Testa configuração CORS"""
    print("\n Testando configuração CORS...")

    try:
        # Fazer uma requisição OPTIONS (preflight)
        response = requests.options(
            "http://localhost:5000/empreendimentos/upload",
            headers={
                'Origin': 'http://localhost:3000',
                'Access-Control-Request-Method': 'POST',
                'Access-Control-Request-Headers': 'Content-Type'
            },
            timeout=10
        )

        cors_headers = ['access-control-allow-origin', 'access-control-allow-methods', 'access-control-allow-headers']

        cors_present = any(h in response.headers for h in cors_headers)
        if cors_present:
            print(" CORS configurado corretamente")
            return True
        else:
            print(" Headers CORS não encontrados")
            print("   Headers de resposta:", list(response.headers.keys()))
            return False

    except Exception as e:
        print(f" Erro ao testar CORS: {str(e)}")
        return False

def testar_tamanho_arquivo():
    """Testa limite de tamanho de arquivo"""
    print("\n Testando limite de tamanho de arquivo...")

    try:
        # Criar arquivo grande (mas ainda dentro do limite)
        wb = Workbook()
        ws = wb.active

        # Adicionar muitos dados
        for i in range(1, 1000):
            ws[f'A{i}'] = f'Dado {i}'
            ws[f'B{i}'] = f'Endereço {i}'
            ws[f'C{i}'] = f'Empresa {i}'

        bio = BytesIO()
        wb.save(bio)
        tamanho = len(bio.getvalue()) / 1024 / 1024  # MB
        print(f"   Tamanho do arquivo de teste: {tamanho:.2f} MB")

        bio.seek(0)
        files = {'file': ('teste_grande.xlsx', bio, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        response = requests.post(
            "http://localhost:5000/empreendimentos/upload",
            files=files,
            timeout=60
        )

        if response.status_code == 413:
            print(" Arquivo muito grande (413 - Payload Too Large)")
            return False
        elif response.status_code == 200:
            print(" Arquivo dentro do limite de tamanho")
            return True
        else:
            print(f"  Status inesperado: {response.status_code}")
            return True

    except Exception as e:
        print(f" Erro ao testar tamanho: {str(e)}")
        return False

def testar_banco_dados():
    """Testa conexão com banco de dados"""
    print("\n Testando conexão com banco de dados...")

    try:
        import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app
        app = create_app()

        with app.app_context():
            from src.db_sql import get_db
            with get_db() as db:
                # Testar uma query simples
                result = db.execute("SELECT 1").fetchone()
                if result:
                    print(" Conexão com banco de dados OK")
                    return True
                else:
                    print(" Query de teste falhou")
                    return False

    except Exception as e:
        print(f" Erro na conexão com banco: {str(e)}")
        return False

def testar_logs():
    """Verifica se há logs de erro recentes"""
    print("\n Verificando logs de erro...")

    try:
        log_files = ['logs/flask.log', 'logs/error.log']
        found_errors = False

        for log_file in log_files:
            if os.path.exists(log_file):
                print(f"   Verificando {log_file}...")
                with open(log_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()[-20:]  # Últimas 20 linhas

                    for line in lines:
                        if 'ERROR' in line or 'Exception' in line:
                            print(f"    Erro encontrado: {line.strip()}")
                            found_errors = True

        if not found_errors:
            print(" Nenhum erro recente encontrado nos logs")
            return True
        else:
            return False

    except Exception as e:
        print(f" Erro ao verificar logs: {str(e)}")
        return False

def executar_diagnostico_completo():
    """Executa todos os testes de diagnóstico"""
    print(" INICIANDO DIAGNÓSTICO COMPLETO DO SISTEMA DE UPLOAD")
    print("=" * 60)

    testes = [
        ("Conexão com Servidor", testar_conexao_servidor),
        ("Configuração CORS", testar_cors),
        ("Conexão com Banco", testar_banco_dados),
        ("Limite de Tamanho", testar_tamanho_arquivo),
        ("Endpoint de Upload", testar_endpoint_upload),
        ("Logs de Erro", testar_logs),
    ]

    resultados = {}

    for nome_teste, funcao_teste in testes:
        try:
            resultado = funcao_teste()
            resultados[nome_teste] = resultado
        except Exception as e:
            print(f" Erro ao executar teste '{nome_teste}': {str(e)}")
            resultados[nome_teste] = False

    print("\n" + "=" * 60)
    print(" RESUMO DOS TESTES:")

    todos_passaram = True
    for nome_teste, resultado in resultados.items():
        status = " PASSOU" if resultado else " FALHOU"
        print(f"   {nome_teste}: {status}")
        if not resultado:
            todos_passaram = False

    print("\n" + "=" * 60)
    if todos_passaram:
        print(" TODOS OS TESTES PASSARAM!")
        print("   Se ainda há erro de upload, pode ser:")
        print("   - Problema no frontend (JavaScript/React)")
        print("   - Arquivo do usuário com formato diferente")
        print("   - Problema de rede entre frontend e backend")
    else:
        print(" ALGUNS TESTES FALHARAM!")
        print("   Os testes que falharam precisam ser corrigidos primeiro.")

    return todos_passaram

if __name__ == "__main__":
    sucesso = executar_diagnostico_completo()
    print(f"\nResultado final: {'PASSOU' if sucesso else 'FALHOU'}")
    sys.exit(0 if sucesso else 1)
