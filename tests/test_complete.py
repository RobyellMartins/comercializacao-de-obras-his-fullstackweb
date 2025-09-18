import os
import sys
import io
import json
from openpyxl import Workbook

# Configurar SQLite para teste
os.environ['DATABASE_URL'] = 'sqlite:///test_complete_obras_his.db'

# Importar e testar a aplicação
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app

def test_construtoras_endpoints():
    """Testa endpoints de construtoras"""
    app = create_app()
    
    with app.test_client() as client:
        print(" Testando endpoints de construtoras...")
        
        # Teste 1: Listar construtoras
        print("\n1. Testando listagem de construtoras...")
        response = client.get('/api/construtoras')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            construtoras = response.get_json()
            print(f"   Construtoras encontradas: {len(construtoras)}")
            for const in construtoras:
                print(f"     - {const.get('nome')}")
        else:
            print(f"   Erro: {response.get_json()}")
        
        # Teste 2: Criar construtora
        print("\n2. Testando criação de construtora...")
        nova_construtora = {
            'nome': 'Construtora Teste API',
            'cnpj': '12.345.678/0001-99',
            'telefone': '(61) 9999-8888',
            'email': 'teste@construtoraapi.com',
            'endereco': 'Rua Teste API, 123'
        }
        
        response = client.post('/api/construtoras', 
                             data=json.dumps(nova_construtora),
                             content_type='application/json')
        print(f"   Status: {response.status_code}")
        if response.status_code in [200, 201]:
            construtora_criada = response.get_json()
            print(f"   Construtora criada com ID: {construtora_criada.get('id')}")
            return construtora_criada.get('id')
        else:
            print(f"   Erro: {response.get_json()}")
            return None

def test_unidades_endpoints():
    """Testa endpoints de unidades"""
    app = create_app()
    
    with app.test_client() as client:
        print("\n Testando endpoints de unidades...")
        
        # Primeiro criar um empreendimento para vincular as unidades
        print("\n1. Criando empreendimento para teste de unidades...")
        empreendimento_teste = {
            'nome': 'Residencial Unidades Teste',
            'nome_empresa': 'Empresa Unidades Teste',
            'cep': '70000-999',
            'endereco': 'Rua Unidades, 999',
            'observacao': 'Teste de unidades'
        }
        
        response = client.post('/empreendimentos',
                             data=json.dumps(empreendimento_teste),
                             content_type='application/json')
        
        if response.status_code != 201:
            print(f"   Erro ao criar empreendimento: {response.get_json()}")
            return
        
        empreendimento_id = response.get_json().get('id')
        print(f"   Empreendimento criado com ID: {empreendimento_id}")
        
        # Teste 2: Listar unidades
        print("\n2. Testando listagem de unidades...")
        response = client.get('/api/unidades')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            unidades = response.get_json()
            print(f"   Unidades encontradas: {len(unidades)}")
        else:
            print(f"   Erro: {response.get_json()}")
        
        # Teste 3: Criar unidade
        print("\n3. Testando criação de unidade...")
        nova_unidade = {
            'empreendimento_id': empreendimento_id,
            'numero_unidade': '999',
            'tamanho_m2': 75.5,
            'preco_venda': 250000.00,
            'mecanismo_pagamento': 'financiamento'
        }
        
        response = client.post('/api/unidades',
                             data=json.dumps(nova_unidade),
                             content_type='application/json')
        print(f"   Status: {response.status_code}")
        if response.status_code in [200, 201]:
            unidade_criada = response.get_json()
            print(f"   Unidade criada com ID: {unidade_criada.get('id')}")
            return unidade_criada.get('id')
        else:
            print(f"   Erro: {response.get_json()}")
            return None

def test_filtros_avancados():
    """Testa filtros avançados na listagem"""
    app = create_app()
    
    with app.test_client() as client:
        print("\n Testando filtros avançados...")
        
        # Teste 1: Filtro somente publicadas
        print("\n1. Testando filtro 'somente_publicadas'...")
        response = client.get('/empreendimentos?somente_publicadas=1')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   Empreendimentos publicados: {len(empreendimentos)}")
        
        # Teste 2: Filtro por nome
        print("\n2. Testando filtro por nome...")
        response = client.get('/empreendimentos?nome=Teste')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   Empreendimentos com 'Teste' no nome: {len(empreendimentos)}")
        
        # Teste 3: Filtro por CEP
        print("\n3. Testando filtro por CEP...")
        response = client.get('/empreendimentos?cep=70000')
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            empreendimentos = response.get_json()
            print(f"   Empreendimentos com CEP '70000': {len(empreendimentos)}")

def test_cenarios_erro():
    """Testa cenários de erro"""
    app = create_app()
    
    with app.test_client() as client:
        print("\n Testando cenários de erro...")
        
        # Teste 1: Criar empreendimento sem dados obrigatórios
        print("\n1. Testando criação sem dados obrigatórios...")
        dados_incompletos = {
            'nome': 'Teste Incompleto'
            # Faltando nome_empresa e cep
        }
        
        response = client.post('/empreendimentos',
                             data=json.dumps(dados_incompletos),
                             content_type='application/json')
        print(f"   Status: {response.status_code}")
        print(f"   Resposta: {response.get_json()}")
        
        # Teste 2: Buscar empreendimento inexistente
        print("\n2. Testando busca de empreendimento inexistente...")
        response = client.get('/empreendimentos/99999')
        print(f"   Status: {response.status_code}")
        print(f"   Resposta: {response.get_json()}")
        
        # Teste 3: Upload de arquivo inválido
        print("\n3. Testando upload de arquivo inválido...")
        arquivo_invalido = io.BytesIO(b"conteudo invalido")
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (arquivo_invalido, 'invalido.txt')},
                             content_type='multipart/form-data')
        print(f"   Status: {response.status_code}")
        print(f"   Resposta: {response.get_json()}")

def test_mecanismos_pagamento():
    """Testa validação de mecanismos de pagamento"""
    app = create_app()
    
    with app.test_client() as client:
        print("\n Testando mecanismos de pagamento...")
        
        # Criar empreendimento para teste
        empreendimento_teste = {
            'nome': 'Residencial Pagamento Teste',
            'nome_empresa': 'Empresa Pagamento',
            'cep': '70000-888',
            'endereco': 'Rua Pagamento, 888'
        }
        
        response = client.post('/empreendimentos',
                             data=json.dumps(empreendimento_teste),
                             content_type='application/json')
        
        if response.status_code != 201:
            print(f"   Erro ao criar empreendimento: {response.get_json()}")
            return
        
        empreendimento_id = response.get_json().get('id')
        
        # Testar cada mecanismo de pagamento
        mecanismos = ['financiamento', 'à vista', 'consórcio', 'outros']
        
        for i, mecanismo in enumerate(mecanismos, 1):
            print(f"\n{i}. Testando mecanismo: {mecanismo}")
            
            unidade_teste = {
                'empreendimento_id': empreendimento_id,
                'numero_unidade': f'PAG{i}',
                'tamanho_m2': 50.0,
                'preco_venda': 150000.00,
                'mecanismo_pagamento': mecanismo
            }
            
            response = client.post('/api/unidades',
                                 data=json.dumps(unidade_teste),
                                 content_type='application/json')
            print(f"   Status: {response.status_code}")
            if response.status_code in [200, 201]:
                print(f"    Mecanismo '{mecanismo}' aceito")
            else:
                print(f"    Erro: {response.get_json()}")

def criar_planilha_com_caracteres_especiais():
    """Cria planilha com caracteres especiais para teste unicode"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos
    ws['A1'] = 'nome_empreendimento'
    ws['B1'] = 'nome_empresa'
    ws['C1'] = 'cep'
    ws['D1'] = 'endereco'
    ws['E1'] = 'observacao'
    ws['F1'] = 'numero_unidade'
    ws['G1'] = 'tamanho_m2'
    ws['H1'] = 'preco_venda'
    ws['I1'] = 'mecanismo_pagamento'
    
    # Dados com caracteres especiais
    dados_unicode = [
        ['Residencial São José', 'Construtora Ação & Cia', '70000-123', 'Rua da Conceição, 123', 'Observação com acentos', '101', '50.5', '150000', 'à vista'],
        ['Condomínio Esperança', 'Empreendimentos Coração Ltda', '70000-456', 'Av. São João, 456', 'Próximo à escola', '201', '60.0', '180000', 'financiamento'],
        ['Residencial Vitória', 'Construtora União & Progresso', '70000-789', 'Rua José da Silva, 789', 'Área de lazer completa', '301', '45.0', '120000', 'consórcio'],
    ]
    
    for i, linha in enumerate(dados_unicode, start=2):
        for j, valor in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=valor)
    
    # Salvar em memória
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def test_unicode_processing():
    """Testa processamento de caracteres unicode"""
    app = create_app()
    
    with app.test_client() as client:
        print("\n Testando processamento Unicode...")
        
        # Criar planilha com caracteres especiais
        planilha_unicode = criar_planilha_com_caracteres_especiais()
        
        # Teste 1: Preview com unicode
        print("\n1. Testando preview com caracteres especiais...")
        response = client.post('/empreendimentos/upload/preview',
                             data={'file': (planilha_unicode, 'unicode_test.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            preview_data = response.get_json()
            print(f"   Preview válido: {preview_data.get('valido')}")
            print(f"   Empreendimentos: {preview_data.get('empreendimentos_encontrados')}")
            print(f"   Unidades: {preview_data.get('unidades_encontradas')}")
            
            # Verificar se caracteres especiais foram preservados
            empreendimentos = preview_data.get('empreendimentos', [])
            for emp in empreendimentos:
                print(f"     - {emp.get('nome')} ({emp.get('nome_empresa')})")
        
        # Teste 2: Upload real com unicode
        print("\n2. Testando upload real com caracteres especiais...")
        planilha_unicode_2 = criar_planilha_com_caracteres_especiais()
        
        response = client.post('/empreendimentos/upload',
                             data={'file': (planilha_unicode_2, 'unicode_test.xlsx')},
                             content_type='multipart/form-data')
        
        print(f"   Status: {response.status_code}")
        if response.status_code == 200:
            upload_data = response.get_json()
            print(f"   Empreendimentos processados: {upload_data.get('empreendimentos_processados')}")
            print(f"   Unidades processadas: {upload_data.get('unidades_processadas')}")
            print(f"   Erros: {upload_data.get('erros')}")

def run_complete_tests():
    """Executa todos os testes completos"""
    print(" Iniciando testes completos do Sistema Obras HIS...")
    
    try:
        # Testes de endpoints
        construtora_id = test_construtoras_endpoints()
        unidade_id = test_unidades_endpoints()
        
        # Testes de filtros
        test_filtros_avancados()
        
        # Testes de cenários de erro
        test_cenarios_erro()
        
        # Testes de mecanismos de pagamento
        test_mecanismos_pagamento()
        
        # Testes de unicode
        test_unicode_processing()
        
        print("\n Todos os testes completos foram executados!")
        
    except Exception as e:
        print(f"\n Erro durante os testes: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    run_complete_tests()
