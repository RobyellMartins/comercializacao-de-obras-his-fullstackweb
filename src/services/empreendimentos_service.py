import openpyxl
from typing import List, Dict, Optional
import logging
import unicodedata
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)

class EmpreendimentosService:
    def __init__(self, repository):
        self.repository = repository
    
    def listar_todos(self) -> List[Dict]:
        """Lista todos os empreendimentos"""
        return self.repository.listar_todos()
    
    def obter_por_id(self, id: int) -> Optional[Dict]:
        """Obtém um empreendimento por ID"""
        return self.repository.obter_por_id(id)
    
    def criar(self, dados: Dict) -> Dict:
        """Cria um novo empreendimento"""
        # Validar dados obrigatórios
        campos_obrigatorios = ['nome', 'nome_empresa', 'cep']
        for campo in campos_obrigatorios:
            if not dados.get(campo):
                raise ValueError(f"Campo '{campo}' é obrigatório")

        # Normalizar dados unicode
        dados = self._normalizar_unicode(dados)
        
        return self.repository.criar(dados)
    
    def atualizar(self, id: int, dados: Dict) -> Optional[Dict]:
        """Atualiza um empreendimento existente"""
        # Verificar se existe
        empreendimento_existente = self.repository.obter_por_id(id)
        if not empreendimento_existente:
            return None
        
        # Normalizar dados unicode
        dados = self._normalizar_unicode(dados)
        
        return self.repository.atualizar(id, dados)
    
    def deletar(self, id: int) -> bool:
        """Deleta um empreendimento"""
        # Verificar se existe
        empreendimento_existente = self.repository.obter_por_id(id)
        if not empreendimento_existente:
            return False
        
        return self.repository.deletar(id)
    
    def publicar_empreendimento(self, id: int) -> Optional[Dict]:
        """Publica um empreendimento e define data de expiração"""
        empreendimento = self.repository.obter_por_id(id)
        if not empreendimento:
            return None
        
        agora = datetime.utcnow()
        expira_em = agora + timedelta(days=30)  # 30 dias conforme requisito
        
        dados_publicacao = {
            'publicado_em': agora,
            'expira_em': expira_em,
            'status_publicacao': 'publicado'
        }
        
        return self.repository.atualizar(id, dados_publicacao)
    
    def aguardar_publicacao(self, id: int) -> Optional[Dict]:
        """Marca empreendimento para aguardar publicação"""
        empreendimento = self.repository.obter_por_id(id)
        if not empreendimento:
            return None
        
        dados_aguardo = {
            'status_publicacao': 'aguardando',
            'publicado_em': None,
            'expira_em': None
        }
        
        return self.repository.atualizar(id, dados_aguardo)
    
    def _normalizar_unicode(self, dados: Dict) -> Dict:
        """Normaliza strings unicode nos dados"""
        dados_normalizados = {}
        
        for chave, valor in dados.items():
            if isinstance(valor, str):
                # Normalizar unicode para forma NFC (Canonical Decomposition, followed by Canonical Composition)
                valor_normalizado = unicodedata.normalize('NFC', valor)
                # Remover espaços extras
                valor_normalizado = ' '.join(valor_normalizado.split())
                dados_normalizados[chave] = valor_normalizado
            else:
                dados_normalizados[chave] = valor
        
        return dados_normalizados
    
    def preview_planilha(self, file) -> Dict:
        """Faz preview dos dados da planilha sem salvar no banco"""
        try:
            # Carregar a planilha
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            empreendimentos_preview = []
            unidades_preview = []
            erros = []

            # Assumir que a primeira linha contém os cabeçalhos
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # Mapear colunas esperadas conforme história de usuário - MAPEAMENTO FLEXÍVEL
            mapeamento_colunas = {
                'NomeEmpreendimento': [
                    'nomeempreendimento', 'nome_empreendimento', 'empreendimento', 'nome',
                    'nome do empreendimento', 'projeto', 'residencial', 'condominio'
                ],
                'NomeEmpresa': [
                    'nomeempresa', 'nome_empresa', 'empresa', 'construtora',
                    'nome da empresa', 'incorporadora', 'construção'
                ],
                'CEP': ['cep'],
                'Endereco': [
                    'endereco', 'endereço', 'endereço', 'endereco completo', 
                    'endereço completo', 'local', 'localização', 'address'
                ],
                'Observacao': [
                    'observacao', 'observação', 'observacoes', 'observações',
                    'notas', 'comentarios', 'comentários', 'obs'
                ],
                'ConstrutoraID': ['construtoraid', 'construtora_id'],
                'NumeroUnidade': [
                    'numerounidade', 'numero_unidade', 'unidade', 'numero',
                    'número', 'apt', 'apartamento', 'casa'
                ],
                'TamanhoM2': [
                    'tamanhom2', 'tamanho_m2', 'area', 'tamanho', 'área',
                    'metragem', 'm2', 'm²', 'metros'
                ],
                'PrecoVenda': [
                    'precovenda', 'preco_venda', 'preco', 'valor', 'preço',
                    'preço de venda', 'price', 'custo'
                ],
                'MecanismoPagamento': [
                    'mecanismopagamento', 'mecanismo_pagamento', 'pagamento',
                    'forma_pagamento', 'forma de pagamento', 'forma_pagto'
                ]
            }

            # Encontrar índices das colunas
            indices_colunas = {}
            for campo, possiveis_nomes in mapeamento_colunas.items():
                for i, header in enumerate(headers):
                    if header and header.lower().strip() in [nome.lower() for nome in possiveis_nomes]:
                        indices_colunas[campo] = i
                        break

            # Verificar se colunas obrigatórias estão presentes
            # Se não tem coluna CEP separada, mas tem endereço, tentaremos extrair o CEP do endereço
            if 'NomeEmpreendimento' not in indices_colunas:
                erros.append(f"Coluna obrigatória 'NomeEmpreendimento' não encontrada na planilha")
            
            if 'CEP' not in indices_colunas and 'Endereco' not in indices_colunas:
                erros.append(f"É necessário ter uma coluna 'CEP' ou 'Endereco' (com CEP incluído) na planilha")

            if erros:
                return {
                    'preview': True,
                    'valido': False,
                    'total_linhas': sheet.max_row - 1,
                    'erros': len(erros),
                    'detalhes_erros': erros,
                    'empreendimentos': [],
                    'unidades': []
                }

            # Agrupar linhas por empreendimento
            empreendimentos_grupo = {}

            # Processar cada linha (começando da linha 2)
            for row_num in range(2, sheet.max_row + 1):
                try:
                    row = sheet[row_num]

                    # Extrair dados do empreendimento
                    nome_emp = self._extrair_valor(row, indices_colunas.get('NomeEmpreendimento'))
                    nome_empresa = self._extrair_valor(row, indices_colunas.get('NomeEmpresa'))
                    cep = self._extrair_valor(row, indices_colunas.get('CEP'))
                    endereco_completo = self._extrair_valor(row, indices_colunas.get('Endereco'))

                    # Se não tem CEP separado, tentar extrair do endereço
                    if not cep and endereco_completo:
                        cep = self._extrair_cep_do_endereco(endereco_completo)
                    
                    # Se não tem nome da empresa, usar o nome do empreendimento
                    if not nome_empresa:
                        nome_empresa = nome_emp

                    if not nome_emp:
                        erros.append(f"Linha {row_num}: NomeEmpreendimento é obrigatório")
                        continue
                    
                    if not cep:
                        erros.append(f"Linha {row_num}: CEP é obrigatório (não encontrado no endereço)")
                        continue

                    chave_emp = (nome_emp, cep)

                    if chave_emp not in empreendimentos_grupo:
                        # Limpar endereço removendo o CEP se foi extraído de lá
                        endereco_limpo = endereco_completo
                        if endereco_completo and cep:
                            endereco_limpo = self._limpar_endereco_sem_cep(endereco_completo)
                        
                        dados_emp = {
                            'nome': nome_emp,
                            'nome_empresa': nome_empresa,
                            'cep': cep,
                            'endereco': endereco_limpo,
                            'observacao': self._extrair_valor(row, indices_colunas.get('Observacao')),
                            'construtora_id': self._parse_int(self._extrair_valor(row, indices_colunas.get('ConstrutoraID')))
                        }
                        
                        # Normalizar unicode
                        dados_emp = self._normalizar_unicode(dados_emp)
                        
                        empreendimentos_grupo[chave_emp] = {
                            'dados': dados_emp,
                            'unidades': []
                        }

                    # Extrair dados da unidade se presente
                    numero_unidade = self._extrair_valor(row, indices_colunas.get('NumeroUnidade'))
                    if numero_unidade:
                        unidade = {
                            'numero_unidade': numero_unidade,
                            'tamanho_m2': self._parse_decimal(self._extrair_valor(row, indices_colunas.get('TamanhoM2'))),
                            'preco_venda': self._parse_decimal(self._extrair_valor(row, indices_colunas.get('PrecoVenda'))),
                            'mecanismo_pagamento': self._extrair_valor(row, indices_colunas.get('MecanismoPagamento')) or 'outros'
                        }
                        empreendimentos_grupo[chave_emp]['unidades'].append(unidade)

                except Exception as e:
                    erros.append(f"Linha {row_num}: {str(e)}")

            # Preparar dados para preview
            for (nome, cep), grupo in empreendimentos_grupo.items():
                empreendimentos_preview.append(grupo['dados'])
                unidades_preview.extend(grupo['unidades'])

            return {
                'preview': True,
                'valido': len(erros) == 0,
                'total_linhas': sheet.max_row - 1,
                'empreendimentos_encontrados': len(empreendimentos_preview),
                'unidades_encontradas': len(unidades_preview),
                'erros': len(erros),
                'detalhes_erros': erros,
                'empreendimentos': empreendimentos_preview,
                'unidades': unidades_preview,
                'resumo': {
                    'total_empreendimentos': len(empreendimentos_preview),
                    'total_unidades': len(unidades_preview),
                    'linhas_processadas': sheet.max_row - 1,
                    'linhas_com_erro': len(erros)
                }
            }

        except Exception as e:
            logger.error(f"Erro ao fazer preview da planilha: {str(e)}")
            raise Exception(f"Erro ao processar planilha para preview: {str(e)}")
    
    def processar_planilha(self, file) -> Dict:
        """Processa uma planilha Excel com dados de empreendimentos e unidades"""
        try:
            # Carregar a planilha
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            empreendimentos_processados = []
            unidades_processadas = []
            erros = []

            # Assumir que a primeira linha contém os cabeçalhos
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # Mapear colunas esperadas conforme história de usuário
            mapeamento_colunas = {
                'NomeEmpreendimento': ['nomeempreendimento', 'nome_empreendimento', 'empreendimento'],
                'NomeEmpresa': ['nomeempresa', 'nome_empresa', 'empresa', 'construtora'],
                'CEP': ['cep'],
                'Endereco': ['endereco', 'endereço'],
                'Observacao': ['observacao', 'observação'],
                'ConstrutoraID': ['construtoraid', 'construtora_id'],
                'NumeroUnidade': ['numerounidade', 'numero_unidade'],
                'TamanhoM2': ['tamanhom2', 'tamanho_m2', 'area'],
                'PrecoVenda': ['precovenda', 'preco_venda'],
                'MecanismoPagamento': ['mecanismopagamento', 'mecanismo_pagamento']
            }

            # Encontrar índices das colunas
            indices_colunas = {}
            for campo, possiveis_nomes in mapeamento_colunas.items():
                for i, header in enumerate(headers):
                    if header and header.lower().strip() in [nome.lower() for nome in possiveis_nomes]:
                        indices_colunas[campo] = i
                        break

            # Verificar se colunas obrigatórias estão presentes
            # Se não tem coluna CEP separada, mas tem endereço, tentaremos extrair o CEP do endereço
            if 'NomeEmpreendimento' not in indices_colunas:
                erros.append(f"Coluna obrigatória 'NomeEmpreendimento' não encontrada na planilha")

            if 'CEP' not in indices_colunas and 'Endereco' not in indices_colunas:
                erros.append(f"É necessário ter uma coluna 'CEP' ou 'Endereco' (com CEP incluído) na planilha")

            if erros:
                return {
                    'processados': 0,
                    'erros': len(erros),
                    'detalhes_erros': erros,
                    'empreendimentos': [],
                    'unidades': []
                }

            # Agrupar linhas por empreendimento
            empreendimentos_grupo = {}

            # Processar cada linha (começando da linha 2)
            for row_num in range(2, sheet.max_row + 1):
                try:
                    row = sheet[row_num]

                    # Extrair dados do empreendimento
                    nome_emp = self._extrair_valor(row, indices_colunas.get('NomeEmpreendimento'))
                    nome_empresa = self._extrair_valor(row, indices_colunas.get('NomeEmpresa'))
                    cep = self._extrair_valor(row, indices_colunas.get('CEP'))
                    endereco_completo = self._extrair_valor(row, indices_colunas.get('Endereco'))

                    # Se não tem CEP separado, tentar extrair do endereço
                    if not cep and endereco_completo:
                        cep = self._extrair_cep_do_endereco(endereco_completo)

                    # Se não tem nome da empresa, usar o nome do empreendimento
                    if not nome_empresa:
                        nome_empresa = nome_emp

                    if not nome_emp:
                        erros.append(f"Linha {row_num}: NomeEmpreendimento é obrigatório")
                        continue

                    if not cep:
                        erros.append(f"Linha {row_num}: CEP é obrigatório (não encontrado no endereço)")
                        continue

                    chave_emp = (nome_emp, cep)

                    if chave_emp not in empreendimentos_grupo:
                        # Limpar endereço removendo o CEP se foi extraído de lá
                        endereco_limpo = endereco_completo
                        if endereco_completo and cep:
                            endereco_limpo = self._limpar_endereco_sem_cep(endereco_completo)

                        dados_emp = {
                            'nome': nome_emp,
                            'nome_empresa': nome_empresa,
                            'cep': cep,
                            'endereco': endereco_limpo,
                            'observacao': self._extrair_valor(row, indices_colunas.get('Observacao')),
                            'construtora_id': self._parse_int(self._extrair_valor(row, indices_colunas.get('ConstrutoraID')))
                        }

                        # Normalizar unicode
                        dados_emp = self._normalizar_unicode(dados_emp)

                        empreendimentos_grupo[chave_emp] = {
                            'dados': dados_emp,
                            'unidades': []
                        }

                    # Extrair dados da unidade se presente
                    numero_unidade = self._extrair_valor(row, indices_colunas.get('NumeroUnidade'))
                    if numero_unidade:
                        unidade = {
                            'numero_unidade': numero_unidade,
                            'tamanho_m2': self._parse_decimal(self._extrair_valor(row, indices_colunas.get('TamanhoM2'))),
                            'preco_venda': self._parse_decimal(self._extrair_valor(row, indices_colunas.get('PrecoVenda'))),
                            'mecanismo_pagamento': self._extrair_valor(row, indices_colunas.get('MecanismoPagamento')) or 'outros'
                        }
                        empreendimentos_grupo[chave_emp]['unidades'].append(unidade)

                except Exception as e:
                    erros.append(f"Linha {row_num}: {str(e)}")

            # Processar os grupos
            for (nome, cep), grupo in empreendimentos_grupo.items():
                try:
                    # Criar empreendimento
                    emp_dados = grupo['dados']
                    empreendimento_criado = self.criar(emp_dados)
                    empreendimentos_processados.append(empreendimento_criado)

                    # Criar unidades
                    for unidade_dados in grupo['unidades']:
                        unidade_dados['empreendimento_id'] = empreendimento_criado['id']
                        # Usar o repository para criar unidade
                        unidade_criada = self.repository.criar_unidade(unidade_dados)
                        unidades_processadas.append(unidade_criada)

                except Exception as e:
                    erros.append(f"Erro ao processar empreendimento {nome}: {str(e)}")

            return {
                'empreendimentos_processados': len(empreendimentos_processados),
                'unidades_processadas': len(unidades_processadas),
                'erros': len(erros),
                'detalhes_erros': erros,
                'empreendimentos': empreendimentos_processados,
                'unidades': unidades_processadas
            }

        except Exception as e:
            logger.error(f"Erro ao processar planilha: {str(e)}")
            raise Exception(f"Erro ao processar planilha: {str(e)}")

    def _extrair_valor(self, row, indice):
        """Extrai valor de uma célula de forma segura"""
        if indice is not None and indice < len(row):
            valor = row[indice].value
            return str(valor).strip() if valor is not None else None
        return None
    
    def _extrair_cep_do_endereco(self, endereco):
        """Extrai CEP do endereço usando regex"""
        if not endereco:
            return None
        
        import re
        # Padrões de CEP: 12345-678, 12345678, cep 12345678, etc.
        padroes_cep = [
            r'cep\s*:?\s*(\d{5}-?\d{3})',  # "cep 12345678" ou "cep: 12345-678"
            r'(\d{5}-\d{3})',              # "12345-678"
            r'(\d{8})',                    # "12345678"
        ]
        
        endereco_lower = endereco.lower()
        
        for padrao in padroes_cep:
            match = re.search(padrao, endereco_lower)
            if match:
                cep = match.group(1)
                # Normalizar formato do CEP
                cep_numeros = re.sub(r'\D', '', cep)
                if len(cep_numeros) == 8:
                    return f"{cep_numeros[:5]}-{cep_numeros[5:]}"
        
        return None
    
    def _limpar_endereco_sem_cep(self, endereco):
        """Remove CEP do endereço deixando apenas o endereço limpo"""
        if not endereco:
            return endereco
        
        import re
        # Remover padrões de CEP do endereço
        padroes_remover = [
            r'\s*-?\s*cep\s*:?\s*\d{5}-?\d{3}',
            r'\s*-?\s*\d{5}-\d{3}',
            r'\s*-?\s*\d{8}',
        ]
        
        endereco_limpo = endereco
        for padrao in padroes_remover:
            endereco_limpo = re.sub(padrao, '', endereco_limpo, flags=re.IGNORECASE)
        
        # Limpar espaços extras e hífens no final
        endereco_limpo = re.sub(r'\s*-\s*$', '', endereco_limpo.strip())
        
        return endereco_limpo.strip()

    def _parse_decimal(self, valor):
        """Converte string para decimal de forma segura"""
        if not valor:
            return None
        try:
            return float(str(valor).replace(',', '.'))
        except (ValueError, TypeError):
            return None
    
    def _parse_int(self, valor):
        """Converte string para inteiro de forma segura"""
        if not valor:
            return None
        try:
            return int(float(str(valor)))
        except (ValueError, TypeError):
            return None
